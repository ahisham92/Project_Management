"""Exporting the setup to a workbook and importing it back."""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.db import connect
from app.excel import ImportError_, read_workbook

from .test_web import text, unlock


@pytest.fixture()
def workbook(signed_in):
    """The project's setup, exported."""
    unlock(signed_in)
    response = signed_in.get("/projects/1/setup/export")
    assert response.status_code == 200
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return openpyxl.load_workbook(io.BytesIO(response.data))


def upload(client, wb, filename="setup.xlsx"):
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return client.post(
        "/projects/1/setup/import",
        data={"workbook": (stream, filename)},
        content_type="multipart/form-data",
        follow_redirects=True,
    )


# --- export -----------------------------------------------------------------

def test_the_export_carries_the_whole_setup(workbook):
    assert workbook.sheetnames == ["Project", "Workflow", "Trades", "Sections", "Deliverables"]
    assert workbook["Deliverables"].max_row == 56, "a header plus 55 deliverables"
    assert workbook["Trades"].max_row == 5
    assert workbook["Sections"].max_row == 6


def test_the_export_writes_dates_as_day_month_year(workbook):
    ws = workbook["Deliverables"]
    header = [c.value for c in ws[1]]
    start = header.index("Start date")
    assert ws.cell(row=2, column=start + 1).value == "31/08/2026"
    assert ws.cell(row=2, column=start + 2).value == "30/09/2026"


def test_the_export_carries_the_workflow_and_the_trade_split(workbook):
    steps = [(r[0], r[1], r[2], r[3]) for r in workbook["Workflow"].iter_rows(min_row=2, values_only=True) if r[0]]
    assert ("Design started", 10, "start", 0) in steps
    assert ("IDC provided", 40, "submission", -5) in steps
    assert ("Code A received", 100, "submission", 14) in steps

    ws = workbook["Deliverables"]
    header = [c.value for c in ws[1]]
    assert "Marine %" in header and "Split total %" in header
    total = header.index("Split total %")
    assert ws.cell(row=2, column=total + 1).value == 100


# --- import -----------------------------------------------------------------

def test_an_edited_workbook_comes_back_in(signed_in, database, workbook):
    deliverables = workbook["Deliverables"]
    header = [c.value for c in deliverables[1]]
    deliverables.cell(row=2, column=header.index("Weight points") + 1).value = 5.0
    deliverables.cell(row=2, column=header.index("Submission date") + 1).value = "15/10/2026"
    workbook["Trades"].cell(row=2, column=2).value = 1200
    workbook["Workflow"].cell(row=3, column=4).value = -7      # IDC now 7 days before submission

    response = upload(signed_in, workbook)
    assert "Imported 55 deliverables" in text(response)

    conn = connect(database)
    try:
        task = conn.execute("SELECT * FROM tasks WHERE wbs = '1.1'").fetchone()
        assert task["weight_points"] == 5.0
        assert task["submission_date"] == "2026-10-15"
        assert conn.execute("SELECT budget_hours FROM trades WHERE key = 'marine'").fetchone()[0] == 1200
        assert conn.execute(
            "SELECT offset_days FROM workflow_steps WHERE name = 'IDC provided'"
        ).fetchone()[0] == -7
        # Nothing was duplicated, and the trade split survived.
        assert conn.execute("SELECT COUNT(*) FROM tasks WHERE project_id = 1").fetchone()[0] == 55
        assert conn.execute(
            "SELECT COUNT(*) FROM task_allocations WHERE task_id = ?", (task["id"],)
        ).fetchone()[0] == 4
    finally:
        conn.close()


def test_import_never_overwrites_progress_reported_since_the_export(signed_in, database, workbook):
    task_id = connect(database).execute("SELECT id FROM tasks WHERE wbs = '2.3'").fetchone()["id"]
    signed_in.post(
        f"/projects/1/tasks/{task_id}/progress",
        data={"status_key": "submitted", "data_date": "01/09/2026"}, follow_redirects=True,
    )
    # The workbook was exported before that status was set, so its Status column
    # is stale — importing it must not roll the deliverable back.
    upload(signed_in, workbook)

    conn = connect(database)
    try:
        row = conn.execute("SELECT * FROM tasks WHERE wbs = '2.3'").fetchone()
        assert row["status_key"] == "submitted"
        assert row["actual_pct"] == 0.8
    finally:
        conn.close()


def test_editing_a_step_percentage_re_derives_progress(signed_in, database, workbook):
    task_id = connect(database).execute("SELECT id FROM tasks WHERE wbs = '2.3'").fetchone()["id"]
    signed_in.post(
        f"/projects/1/tasks/{task_id}/progress",
        data={"status_key": "idc", "data_date": "01/09/2026"}, follow_redirects=True,
    )
    workbook["Workflow"].cell(row=3, column=2).value = 45      # IDC worth 45% instead of 40%
    upload(signed_in, workbook)

    conn = connect(database)
    try:
        assert conn.execute("SELECT actual_pct FROM tasks WHERE id = ?", (task_id,)).fetchone()[0] == 0.45
    finally:
        conn.close()


def test_a_deliverable_removed_from_the_workbook_is_removed(signed_in, database, workbook):
    workbook["Deliverables"].delete_rows(2)
    response = upload(signed_in, workbook)
    assert "Imported 54 deliverables" in text(response)

    conn = connect(database)
    try:
        assert conn.execute("SELECT COUNT(*) FROM tasks WHERE project_id = 1").fetchone()[0] == 54
        assert conn.execute("SELECT 1 FROM tasks WHERE wbs = '1.1'").fetchone() is None
    finally:
        conn.close()


def test_a_split_that_does_not_total_100_is_refused(signed_in, database, workbook):
    ws = workbook["Deliverables"]
    header = [c.value for c in ws[1]]
    ws.cell(row=2, column=header.index("Marine %") + 1).value = 5      # was 35
    response = upload(signed_in, workbook)
    assert "totals 70.0%, not 100%" in text(response)

    conn = connect(database)
    try:
        # Nothing was written: the file is parsed in full before anything changes.
        assert conn.execute("SELECT weight_points FROM tasks WHERE wbs = '1.1'").fetchone()[0] == 2.3
    finally:
        conn.close()


def test_a_row_without_a_date_is_refused(signed_in, workbook):
    ws = workbook["Deliverables"]
    header = [c.value for c in ws[1]]
    ws.cell(row=2, column=header.index("Start date") + 1).value = None
    ws.cell(row=2, column=header.index("Submission date") + 1).value = None
    assert "has no start or submission date" in text(upload(signed_in, workbook))


def test_a_file_that_is_not_a_workbook_is_refused(signed_in):
    unlock(signed_in)
    response = signed_in.post(
        "/projects/1/setup/import",
        data={"workbook": (io.BytesIO(b"not a spreadsheet"), "notes.txt")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert "expects an .xlsx file" in text(response)


def test_importing_while_locked_is_refused(signed_in, workbook):
    signed_in.post("/projects/1/setup/unlock", data={"action": "lock"})
    assert "setup sheet is locked" in text(upload(signed_in, workbook))


def test_reading_a_workbook_without_the_expected_sheets_is_reported():
    wb = openpyxl.Workbook()
    wb.active.title = "Something else"
    stream = io.BytesIO()
    wb.save(stream)
    with pytest.raises(ImportError_, match="missing the Deliverables and Trades sheet"):
        read_workbook(stream.getvalue())
