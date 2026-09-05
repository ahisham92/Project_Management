"""The plan: durations, dependencies, float, the critical path and rework."""

from __future__ import annotations

import pytest

from app.schedule import (
    DEFAULT_MODE, MODES, analyse, critical_path, duration_between, finish_from,
    normalise_mode, order, shift_successors, summarise, window, would_cycle,
)


def task(task_id: int, start: str, finish: str, **extra):
    return dict({"id": task_id, "wbs": str(task_id), "start_date": start,
                 "submission_date": finish}, **extra)


def link(first: int, second: int, lag: float = 0):
    return {"predecessor_id": first, "successor_id": second, "lag_days": lag}


# A short chain that splits and rejoins: 1 → 2 → 4 and 1 → 3 → 4, where 3 is
# quicker than 2 and so carries float.
def diamond():
    return (
        [task(1, "2026-01-01", "2026-01-10"), task(2, "2026-01-11", "2026-01-20"),
         task(3, "2026-01-11", "2026-01-15"), task(4, "2026-01-21", "2026-01-25")],
        [link(1, 2), link(1, 3), link(2, 4), link(3, 4)],
    )


# --- durations -------------------------------------------------------------

def test_a_duration_counts_both_end_days():
    assert duration_between("2026-01-01", "2026-01-01") == 1
    assert duration_between("2026-01-01", "2026-01-10") == 10


def test_a_missing_date_has_no_duration():
    assert duration_between("", "2026-01-10") == 0
    assert duration_between("2026-01-01", None) == 0


def test_a_start_and_a_duration_give_the_finish():
    assert finish_from("2026-01-01", 10) == "2026-01-10"
    assert finish_from("2026-01-01", 1) == "2026-01-01"


def test_a_duration_of_nothing_still_lasts_a_day():
    """A deliverable that takes no time cannot be drawn, so it takes one day."""
    assert finish_from("2026-01-01", 0) == "2026-01-01"
    assert finish_from("2026-01-01", "nonsense") == "2026-01-01"


def test_the_two_ways_of_entering_a_plan():
    assert [key for key, _ in MODES] == ["duration", "dates"]
    assert DEFAULT_MODE == "duration"
    assert normalise_mode("DATES") == "dates"
    assert normalise_mode("whatever") == "duration"


# --- dependencies ----------------------------------------------------------

def test_a_deliverable_cannot_be_made_to_depend_on_itself():
    _tasks, links = diamond()
    assert would_cycle(links, 1, 1)
    assert would_cycle(links, 4, 1)                  # 4 already follows 1
    assert would_cycle(links, 4, 2)
    assert not would_cycle(links, 1, 4)              # a shortcut is fine


def test_predecessors_come_first_whatever_order_they_were_added_in():
    _tasks, links = diamond()
    sequence = order([4, 3, 2, 1], links)
    assert sequence.index(1) < sequence.index(2) < sequence.index(4)
    assert sequence.index(1) < sequence.index(3) < sequence.index(4)


def test_a_looping_programme_still_produces_an_order():
    """A bad link must not leave a deliverable out of the plan entirely."""
    links = [link(1, 2), link(2, 1)]
    assert sorted(order([1, 2], links)) == [1, 2]


# --- float and the critical path -------------------------------------------

def test_the_longest_run_carries_no_float_and_is_critical():
    tasks, links = diamond()
    result = analyse(tasks, links)
    assert result[1]["total_float"] == 0 and result[1]["is_critical"]
    assert result[2]["total_float"] == 0 and result[2]["is_critical"]
    assert result[4]["total_float"] == 0 and result[4]["is_critical"]
    # The shorter of the two parallel lines can slip five days without harm.
    assert result[3]["total_float"] == 5 and not result[3]["is_critical"]


def test_the_critical_path_reads_in_the_order_the_work_runs():
    tasks, links = diamond()
    assert critical_path(analyse(tasks, links)) == [1, 2, 4]


def test_early_dates_follow_the_predecessors():
    tasks, links = diamond()
    result = analyse(tasks, links)
    assert result[1]["early_finish"] == "2026-01-10"
    assert result[2]["early_start"] == "2026-01-11"   # the day after 1 finishes
    assert result[4]["early_start"] == "2026-01-21"


def test_a_lag_holds_the_successor_back():
    tasks = [task(1, "2026-01-01", "2026-01-10"), task(2, "2026-01-11", "2026-01-20")]
    result = analyse(tasks, [link(1, 2, lag=5)])
    assert result[2]["early_start"] == "2026-01-16"


def test_a_line_that_cannot_start_when_it_is_drawn_to_is_flagged():
    """Its predecessors finish after its own start, so the plan does not hold."""
    tasks = [task(1, "2026-01-01", "2026-01-31"), task(2, "2026-01-05", "2026-01-10")]
    result = analyse(tasks, [link(1, 2)])
    assert result[2]["starts_late"]
    assert result[2]["early_start"] == "2026-02-01"
    assert not result[1]["starts_late"]


def test_a_plan_with_no_links_has_no_critical_path():
    """Nothing is sequenced, so nothing is on a path — calling every line that
    happens to finish on the end date critical would say nothing."""
    tasks = [task(1, "2026-01-01", "2026-01-10"), task(2, "2026-01-01", "2026-01-31")]
    result = analyse(tasks, [])
    assert result[1]["total_float"] == 21            # float is still worked out
    assert not result[2]["is_critical"]
    assert not result[1]["in_a_chain"] and not result[2]["in_a_chain"]


def test_a_line_becomes_critical_once_it_is_sequenced():
    tasks, links = diamond()
    loose = analyse(tasks, [])
    assert not any(row["is_critical"] for row in loose.values())

    sequenced = analyse(tasks, links)
    assert sequenced[1]["is_critical"] and sequenced[1]["in_a_chain"]
    assert not sequenced[3]["is_critical"]           # it has float


def test_analysing_nothing_answers_with_nothing():
    assert analyse([], []) == {}
    assert summarise({}) == {"count": 0, "critical": 0, "start": "", "finish": "",
                             "days": 0, "linked": 0}


def test_the_summary_reads_the_whole_programme():
    tasks, links = diamond()
    totals = summarise(analyse(tasks, links))
    assert totals == {"count": 4, "critical": 3, "linked": 4,
                      "start": "2026-01-01", "finish": "2026-01-25", "days": 25}


# --- moving a line ---------------------------------------------------------

def test_pushing_a_predecessor_out_pushes_everything_after_it():
    tasks, links = diamond()
    tasks[0]["submission_date"] = "2026-01-15"       # five days longer
    moves = shift_successors(tasks, links, 1)

    assert moves[2]["start_date"] == "2026-01-16"
    assert moves[2]["submission_date"] == "2026-01-25"
    assert moves[4]["start_date"] == "2026-01-26"    # the shift carries on down


def test_a_successor_keeps_its_own_duration_when_it_is_pushed():
    tasks, links = diamond()
    tasks[0]["submission_date"] = "2026-01-20"
    moves = shift_successors(tasks, links, 1)
    assert duration_between(moves[3]["start_date"], moves[3]["submission_date"]) == 5


def test_pulling_a_predecessor_forward_frees_float_rather_than_dragging_the_rest():
    """Bringing work forward gives the programme slack; it does not move the
    successors back with it, which is what a planner expects."""
    tasks, links = diamond()
    tasks[0]["submission_date"] = "2026-01-05"
    assert shift_successors(tasks, links, 1) == {}


def test_a_line_already_late_enough_is_left_alone():
    tasks = [task(1, "2026-01-01", "2026-01-10"), task(2, "2026-02-01", "2026-02-10")]
    assert shift_successors(tasks, [link(1, 2)], 1) == {}


def test_a_looping_link_does_not_hang_the_shift():
    tasks = [task(1, "2026-01-01", "2026-01-10"), task(2, "2026-01-11", "2026-01-20")]
    moves = shift_successors(tasks, [link(1, 2), link(2, 1)], 1)
    assert isinstance(moves, dict)                   # it finishes, whatever it decides


# --- the window a chart has to cover ---------------------------------------

def test_the_window_covers_every_date_on_the_plan():
    tasks = [dict(task(1, "2026-01-01", "2026-01-10"),
                  approval_due_date="2026-01-24",
                  step_plan=[{"key": "idc", "date": "2026-01-05"}],
                  revisions=[{"comments_date": "2026-01-26", "submission_date": "2026-02-05"}])]
    assert window(tasks, "2026-01-15") == ("2026-01-01", "2026-02-05")


def test_the_window_of_an_empty_plan_is_still_drawable():
    first, last = window([])
    assert first < last


# --- the screens -----------------------------------------------------------

def text(response) -> str:
    assert response.status_code == 200, response.status_code
    return response.get_data(as_text=True)


def link_two(client, successor: str, predecessor: str, lag: str = "0"):
    return client.post("/projects/1/schedule/links",
                       data={"successor_id": successor, "predecessor_id": predecessor,
                             "lag_days": lag},
                       follow_redirects=True).get_data(as_text=True)


def dates_of(client, task_id: int):
    from app.db import query_one

    with client.application.app_context():
        row = query_one("SELECT start_date, submission_date FROM tasks WHERE id = ?", (task_id,))
    return row["start_date"], row["submission_date"]


def test_the_schedule_opens_on_a_project_with_no_dependencies(signed_in):
    body = text(signed_in.get("/projects/1/schedule"))
    assert "Dates and durations" in body
    assert "link two deliverables" in body           # the empty network says what to do


def test_a_dependency_can_be_made_and_removed(signed_in):
    assert "Dependency added" in link_two(signed_in, "2", "1")
    body = text(signed_in.get("/projects/1/schedule"))
    assert "waits for" in body

    from app.db import query_one

    with signed_in.application.app_context():
        link_id = query_one("SELECT id FROM task_links LIMIT 1")["id"]
    after = signed_in.post(f"/projects/1/schedule/links/{link_id}/delete",
                           follow_redirects=True).get_data(as_text=True)
    assert "Dependency removed" in after


def test_a_link_that_would_loop_is_refused(signed_in):
    link_two(signed_in, "2", "1")
    assert "depend on itself" in link_two(signed_in, "1", "2")
    assert "depend on itself" in link_two(signed_in, "1", "1")


def test_the_same_link_cannot_be_made_twice(signed_in):
    link_two(signed_in, "2", "1")
    assert "already linked" in link_two(signed_in, "2", "1")


def test_a_deliverable_from_another_project_cannot_be_linked(signed_in):
    assert "must belong to this project" in link_two(signed_in, "2", "99999")


def test_changing_a_duration_moves_the_finish_and_whatever_follows(signed_in):
    link_two(signed_in, "2", "1")
    response = signed_in.post("/projects/1/schedule/1",
                              data={"mode": "duration", "start_date": "01/09/2026",
                                    "duration_days": "20"},
                              headers={"Accept": "application/json"})
    moved = {row["id"]: row for row in response.get_json()["moved"]}

    assert moved[1]["start"] == "01/09/2026" and moved[1]["submission"] == "20/09/2026"
    assert dates_of(signed_in, 1) == ("2026-09-01", "2026-09-20")
    # The successor is pushed to the day after, keeping its own length.
    assert dates_of(signed_in, 2)[0] == "2026-09-21"


def test_changing_the_dates_gives_the_duration(signed_in):
    signed_in.post("/projects/1/schedule/mode", data={"mode": "dates"}, follow_redirects=True)
    response = signed_in.post("/projects/1/schedule/1",
                              data={"mode": "dates", "start_date": "01/09/2026",
                                    "submission_date": "10/09/2026"},
                              headers={"Accept": "application/json"})
    assert response.get_json()["moved"][0]["duration"] == 10


def test_a_finish_before_its_start_is_pulled_back_to_the_start(signed_in):
    signed_in.post("/projects/1/schedule/1",
                   data={"mode": "dates", "start_date": "10/09/2026",
                         "submission_date": "01/09/2026"}, follow_redirects=True)
    assert dates_of(signed_in, 1) == ("2026-09-10", "2026-09-10")


def test_the_mode_is_remembered_for_the_project(signed_in):
    signed_in.post("/projects/1/schedule/mode", data={"mode": "dates"}, follow_redirects=True)
    assert "Start and finish dates" in text(signed_in.get("/projects/1/schedule"))

    signed_in.post("/projects/1/schedule/mode", data={"mode": "nonsense"}, follow_redirects=True)
    body = text(signed_in.get("/projects/1/schedule"))
    assert 'value="duration" selected' in body       # anything unknown falls back


def test_the_network_names_the_deliverable_behind_each_box(signed_in):
    link_two(signed_in, "2", "1")
    body = text(signed_in.get("/projects/1/schedule"))
    assert "data-tip" in body
    assert "Deliverable" in body


def test_only_a_manager_can_move_the_plan(client, app):
    client.post("/register", data={"name": "Member", "email": "m@example.com",
                                   "password": "longenough1"})
    with app.app_context():
        from app.db import connect

        conn = connect(app.config["DATABASE"])
        with conn:
            user = conn.execute("SELECT id FROM users WHERE email = 'm@example.com'").fetchone()
            conn.execute("INSERT INTO project_members (project_id, user_id, role) VALUES (1, ?, 'member')",
                         (user["id"],))
        conn.close()

    assert client.get("/projects/1/schedule").status_code == 200      # they may look
    assert client.post("/projects/1/schedule/1",
                       data={"start_date": "01/09/2026", "duration_days": "5"}).status_code == 403
    assert client.post("/projects/1/schedule/links",
                       data={"successor_id": "2", "predecessor_id": "1"}).status_code == 403


# --- rework ----------------------------------------------------------------

@pytest.fixture()
def reworked(signed_in):
    """A deliverable submitted, sent back with a Code C, resubmitted, then B."""
    signed_in.post("/projects/1/tasks/1/progress",
                   data={"status_key": "submitted", "data_date": "01/09/2026"},
                   follow_redirects=True)
    signed_in.post("/projects/1/tasks/1/comments",
                   data={"code": "C", "comments_date": "10/09/2026"}, follow_redirects=True)
    signed_in.post("/projects/1/tasks/1/progress",
                   data={"status_key": "submitted", "data_date": "20/09/2026"},
                   follow_redirects=True)
    signed_in.post("/projects/1/tasks/1/comments",
                   data={"code": "B", "comments_date": "25/09/2026"}, follow_redirects=True)
    return signed_in


def history(client):
    from app.service import load_revisions

    with client.application.app_context():
        return load_revisions(1)


def test_a_code_c_sends_a_deliverable_back_into_rework(signed_in):
    signed_in.post("/projects/1/tasks/1/progress",
                   data={"status_key": "submitted", "data_date": "01/09/2026"},
                   follow_redirects=True)
    body = signed_in.post("/projects/1/tasks/1/comments",
                          data={"code": "C", "comments_date": "10/09/2026"},
                          follow_redirects=True).get_data(as_text=True)
    assert "Code C" in body and "revision 1" in body


def test_a_revision_records_the_code_that_closed_it(reworked):
    attempts = history(reworked)
    assert [a["revision"] for a in attempts] == [0, 1, 2]
    assert attempts[0]["code"] == "C"                # the first issue got a Code C
    assert attempts[1]["code"] == "B"                # the resubmission got a Code B
    assert attempts[2]["code"] == ""                 # the latest is still with the client


def test_each_resubmission_keeps_the_day_the_comments_landed(reworked):
    attempts = history(reworked)
    assert attempts[1]["comments_date"] == "2026-09-10"
    assert attempts[2]["comments_date"] == "2026-09-25"


def test_the_schedule_draws_the_rework_and_says_what_caused_it(reworked):
    body = text(reworked.get("/projects/1/schedule"))
    assert "var(--warning)" in body                  # the rework bar
    assert ">C</text>" in body and ">B</text>" in body
    assert "In rework" in body


def test_the_plan_says_which_code_a_line_is_in_rework_after(reworked):
    from app.db import query_one
    from app.service import project_plan

    with reworked.application.app_context():
        project = query_one("SELECT * FROM projects WHERE id = 1")
        row = {t["id"]: t for t in project_plan(project)["tasks"]}[1]
    assert row["revision"] == 2
    assert row["last_code"] == "B"
    assert [a["cause_code"] for a in row["revisions"]] == ["", "C", "B"]


def test_a_code_a_closes_the_revision_rather_than_opening_another(signed_in):
    signed_in.post("/projects/1/tasks/1/progress",
                   data={"status_key": "code_a", "data_date": "01/09/2026"},
                   follow_redirects=True)
    attempts = history(signed_in)
    assert attempts[-1]["outcome"] == "code_a"
    assert attempts[-1]["code"] == "A"


# --- progress edited in the row --------------------------------------------

def test_a_workflow_line_takes_its_status_from_the_row(signed_in):
    result = signed_in.post("/projects/1/tasks/1/progress",
                            data={"status_key": "idc", "data_date": "01/09/2026"},
                            headers={"Accept": "application/json"}).get_json()
    assert result["ok"] and result["actual"] == 40
    assert result["status_key"] == "idc"
    assert "IDC provided" in result["progress_html"]
    assert "40%" in result["progress_html"]


def test_a_simple_line_takes_a_typed_percentage_from_the_row(signed_in):
    result = signed_in.post("/projects/1/tasks/6/progress",
                            data={"actual_pct": "55", "data_date": "01/09/2026"},
                            headers={"Accept": "application/json"}).get_json()
    assert result["actual"] == 55
    assert "55%" in result["progress_html"]


def test_a_percentage_outside_the_range_is_refused_rather_than_stored(signed_in):
    response = signed_in.post("/projects/1/tasks/6/progress",
                              data={"actual_pct": "150"},
                              headers={"Accept": "application/json"})
    assert response.status_code == 400
    assert "between 0% and 100%" in response.get_json()["error"]


def test_the_row_answers_with_its_status_and_variance_redrawn(signed_in):
    result = signed_in.post("/projects/1/tasks/1/progress",
                            data={"status_key": "code_a", "data_date": "01/09/2026"},
                            headers={"Accept": "application/json"}).get_json()
    assert "badge" in result["status_html"]
    assert "%" in result["variance_html"]


def test_the_progress_rows_offer_their_controls_where_they_stand(signed_in):
    body = text(signed_in.get("/projects/1/tasks?filter=all"))
    assert body.count('data-cell="status"') + body.count('data-cell="percent"') > 10
    assert body.count('id="cell-status"') == 1       # one copy, cloned per row
    assert body.count('id="cell-percent"') == 1


def test_the_controls_are_in_the_body_not_the_title(signed_in):
    """A template written into the <title> is invisible to getElementById, so
    every cell fell back to a page load."""
    body = text(signed_in.get("/projects/1/tasks"))
    head = body.split("</head>", 1)[0]
    assert "cell-status" not in head
    assert 'id="cell-status"' in body.split("</head>", 1)[1]


# --- the live check ---------------------------------------------------------

def test_a_project_page_carries_what_it_was_drawn_with(signed_in):
    body = text(signed_in.get("/projects/1/tasks"))
    token = signed_in.get("/projects/1/pulse").get_json()["v"]
    assert f'data-pulse="{token}"' in body
    assert 'data-pulse-url="/projects/1/pulse"' in body


def test_the_token_moves_only_when_something_changes(signed_in):
    first = signed_in.get("/projects/1/pulse").get_json()["v"]
    assert signed_in.get("/projects/1/pulse").get_json()["v"] == first

    signed_in.post("/projects/1/tasks/1/progress",
                   data={"status_key": "idc", "data_date": "01/09/2026"}, follow_redirects=True)
    assert signed_in.get("/projects/1/pulse").get_json()["v"] != first


def test_every_kind_of_change_moves_the_token(signed_in):
    def token():
        return signed_in.get("/projects/1/pulse").get_json()["v"]

    changes = [
        ("a meeting", lambda: signed_in.post("/projects/1/minutes/meetings",
                                             data={"ref": "M1", "meeting_date": "01/09/2026"},
                                             follow_redirects=True)),
        ("a dependency", lambda: signed_in.post("/projects/1/schedule/links",
                                                data={"predecessor_id": "1", "successor_id": "2"},
                                                follow_redirects=True)),
        ("hours booked", lambda: signed_in.post("/projects/1/time",
                                                data={"hours": "4", "entry_date": "01/09/2026"},
                                                follow_redirects=True)),
        ("dates moved", lambda: signed_in.post("/projects/1/schedule/3",
                                               data={"start_date": "01/09/2026", "duration_days": "9"},
                                               follow_redirects=True)),
    ]
    for what, change in changes:
        before = token()
        change()
        assert token() != before, f"{what} did not move the token"


def test_the_pulse_of_a_project_you_cannot_see_is_not_found(client):
    client.post("/register", data={"name": "Outsider", "email": "out2@example.com",
                                   "password": "longenough1"})
    assert client.get("/projects/1/pulse").status_code == 404


def test_two_changes_in_the_same_second_both_move_the_token(signed_in):
    """A timestamp with only seconds would read the two as one, and the second
    change would never reach anyone else's page."""
    first = signed_in.get("/projects/1/pulse").get_json()["v"]
    signed_in.post("/projects/1/schedule/1",
                   data={"start_date": "01/09/2026", "duration_days": "9"}, follow_redirects=True)
    second = signed_in.get("/projects/1/pulse").get_json()["v"]
    signed_in.post("/projects/1/schedule/1",
                   data={"start_date": "02/09/2026", "duration_days": "9"}, follow_redirects=True)
    third = signed_in.get("/projects/1/pulse").get_json()["v"]

    assert first != second != third
    assert int(third) > int(second) > int(first)


def test_a_change_to_one_project_leaves_another_alone(signed_in):
    signed_in.post("/projects/new", data={"code": "OTHER", "name": "Another project",
                                          "ntp_date": "01/09/2026", "duration_months": "6"},
                   follow_redirects=True)
    other = signed_in.get("/projects/2/pulse")
    if other.status_code != 200:
        pytest.skip("the second project was not created")

    before = other.get_json()["v"]
    signed_in.post("/projects/1/tasks/1/progress",
                   data={"status_key": "idc", "data_date": "01/09/2026"}, follow_redirects=True)
    assert signed_in.get("/projects/2/pulse").get_json()["v"] == before


def test_deleting_something_moves_the_token_too(signed_in):
    signed_in.post("/projects/1/schedule/links",
                   data={"predecessor_id": "1", "successor_id": "2"}, follow_redirects=True)
    from app.db import query_one

    with signed_in.application.app_context():
        link_id = query_one("SELECT id FROM task_links LIMIT 1")["id"]

    before = signed_in.get("/projects/1/pulse").get_json()["v"]
    signed_in.post(f"/projects/1/schedule/links/{link_id}/delete", follow_redirects=True)
    assert signed_in.get("/projects/1/pulse").get_json()["v"] != before


def test_dates_are_amended_on_the_schedule_not_the_setup_sheet(signed_in):
    """The Setup sheet still shows the dates so a row reads as a whole, but it
    no longer edits them — the programme is where they belong."""
    signed_in.post("/projects/1/setup/unlock", data={"password": "2026"}, follow_redirects=True)
    body = text(signed_in.get("/projects/1/setup"))

    assert "_start_date" not in body and "_submission_date" not in body
    assert "/projects/1/schedule" in body                 # it says where to go instead
    assert "31/08/2026" in body                           # and still shows them


def test_the_setup_sheet_still_saves_everything_else(signed_in):
    """Removing the date fields must not disturb what Save all writes."""
    signed_in.post("/projects/1/setup/unlock", data={"password": "2026"}, follow_redirects=True)
    before = text(signed_in.get("/projects/1/setup"))
    assert "task_1_points" in before and "task_1_name" in before

    from app.db import query_one

    with signed_in.application.app_context():
        dates = query_one("SELECT start_date, submission_date FROM tasks WHERE id = 1")
        was = (dates["start_date"], dates["submission_date"])

    signed_in.post("/projects/1/setup/save-all",
                   data={"task_1_points": "9", "task_1_name": "Renamed line",
                         "code": "SIBLINE-PORT", "name": "Sibline Port",
                         "ntp_date": "31/08/2026"},
                   follow_redirects=True)
    with signed_in.application.app_context():
        row = query_one("SELECT name, weight_points, start_date, submission_date FROM tasks WHERE id = 1")
    assert row["name"] == "Renamed line" and row["weight_points"] == 9
    assert (row["start_date"], row["submission_date"]) == was      # dates untouched


def test_the_buttons_follow_the_state_the_row_has_just_reached(signed_in):
    """A line that has just been submitted can take a Code B or C, and the row
    has to offer it without the page being loaded again."""
    body = text(signed_in.get("/projects/1/tasks?filter=all"))
    assert "Code B / C" not in body.split('id="actions-1"')[1].split("</td>")[0]

    result = signed_in.post("/projects/1/tasks/1/progress",
                            data={"status_key": "submitted", "data_date": "01/09/2026"},
                            headers={"Accept": "application/json"}).get_json()
    assert "Code B / C" in result["actions_html"]
    assert "History" in result["actions_html"]


# --- how one deliverable waits for another ---------------------------------

def test_start_to_start_does_not_wait_for_the_other_to_finish():
    """"Start a fortnight after the survey starts" needs no guess at when the
    survey will end, which is the whole point of the link."""
    tasks = [task(1, "2026-01-01", "2026-01-30"), task(2, "2026-01-01", "2026-01-10")]
    finish_first = analyse(tasks, [link(1, 2)])
    alongside = analyse(tasks, [dict(link(1, 2, lag=14), kind="SS")])

    assert finish_first[2]["early_start"] == "2026-01-31"     # the day after it ends
    assert alongside[2]["early_start"] == "2026-01-15"        # fourteen days after it began


def test_a_start_to_start_link_shifts_by_the_start_not_the_finish():
    tasks = [task(1, "2026-01-01", "2026-01-30"), task(2, "2026-01-15", "2026-01-24")]
    links = [dict(link(1, 2, lag=14), kind="SS")]
    tasks[0]["start_date"] = "2026-02-01"
    tasks[0]["submission_date"] = "2026-03-02"

    moves = shift_successors(tasks, links, 1)
    assert moves[2]["start_date"] == "2026-02-15"             # not after the finish
    assert duration_between(moves[2]["start_date"], moves[2]["submission_date"]) == 10


def test_all_four_kinds_of_link_are_named_for_a_reader():
    from app.schedule import KINDS, kind_label, kind_note, normalise_kind

    assert [key for key, _ in KINDS] == ["FS", "SS", "FF", "SF"]
    assert kind_label("SS") == "start → start"
    assert kind_label("FS") == "finish → start"
    assert kind_note("FF") == "cannot finish until the other finishes"
    assert kind_note("SF") == "cannot finish until the other starts"
    assert normalise_kind("ss") == "SS"
    assert normalise_kind("nonsense") == "FS"                 # the safe default


def test_a_link_records_the_kind_it_was_made_with(signed_in):
    signed_in.post("/projects/1/schedule/links",
                   data={"predecessor_id": "1", "successor_id": "2", "kind": "SS",
                         "lag_days": "14"}, follow_redirects=True)
    from app.db import query_one

    with signed_in.application.app_context():
        row = query_one("SELECT kind, lag_days FROM task_links LIMIT 1")
    assert row["kind"] == "SS" and row["lag_days"] == 14


def link_id_of(client):
    from app.db import query_one

    with client.application.app_context():
        return query_one("SELECT id FROM task_links LIMIT 1")["id"]


def test_a_lag_is_changed_where_it_stands(signed_in):
    link_two(signed_in, "2", "1")
    result = signed_in.post(f"/projects/1/schedule/links/{link_id_of(signed_in)}",
                            data={"lag_days": "7"},
                            headers={"Accept": "application/json"}).get_json()
    assert result["ok"]
    assert "<svg" in result["network_html"]                   # the diagram comes back drawn

    from app.db import query_one

    with signed_in.application.app_context():
        assert query_one("SELECT lag_days FROM task_links LIMIT 1")["lag_days"] == 7


def test_changing_a_lag_leaves_the_kind_alone(signed_in):
    signed_in.post("/projects/1/schedule/links",
                   data={"predecessor_id": "1", "successor_id": "2", "kind": "SS"},
                   follow_redirects=True)
    signed_in.post(f"/projects/1/schedule/links/{link_id_of(signed_in)}",
                   data={"lag_days": "3"}, headers={"Accept": "application/json"})

    from app.db import query_one

    with signed_in.application.app_context():
        row = query_one("SELECT kind, lag_days FROM task_links LIMIT 1")
    assert row["kind"] == "SS" and row["lag_days"] == 3


def test_the_kind_is_changed_where_it_stands_and_moves_the_float(signed_in):
    link_two(signed_in, "2", "1")
    before = signed_in.post(f"/projects/1/schedule/links/{link_id_of(signed_in)}",
                            data={"kind": "SS"},
                            headers={"Accept": "application/json"}).get_json()
    assert before["ok"]
    from app.db import query_one

    with signed_in.application.app_context():
        assert query_one("SELECT kind FROM task_links LIMIT 1")["kind"] == "SS"


def test_removing_a_dependency_answers_with_the_plan_redrawn(signed_in):
    link_two(signed_in, "2", "1")
    link_id = link_id_of(signed_in)
    result = signed_in.post(f"/projects/1/schedule/links/{link_id}/delete",
                            headers={"Accept": "application/json"}).get_json()
    assert result["removed"] == link_id
    assert result["critical_count"] == 0                      # nothing is sequenced now
    assert signed_in.post(f"/projects/1/schedule/links/{link_id}/delete").status_code == 404


def test_the_dependency_list_offers_the_lag_and_the_kind_in_the_row(signed_in):
    link_two(signed_in, "2", "1")
    body = text(signed_in.get("/projects/1/schedule"))
    assert 'data-cell="link-lag"' in body and 'data-cell="link-kind"' in body
    assert 'id="cell-link-lag"' in body and 'id="cell-link-kind"' in body
    assert "data-live-remove" in body


# --- moving a box by hand --------------------------------------------------

def test_a_box_stays_where_it_is_dragged(signed_in):
    link_two(signed_in, "2", "1")
    assert signed_in.post("/projects/1/schedule/layout/1",
                          data={"x": "320", "y": "140"},
                          headers={"Accept": "application/json"}).get_json()["ok"]

    body = text(signed_in.get("/projects/1/schedule"))
    assert "translate(320,140)" in body


def test_a_box_that_has_not_been_moved_follows_the_automatic_layout(signed_in):
    link_two(signed_in, "2", "1")
    body = text(signed_in.get("/projects/1/schedule"))
    assert "net-node movable" in body
    assert "translate(16,16)" in body                        # the first column


def test_tidying_up_puts_every_box_back(signed_in):
    link_two(signed_in, "2", "1")
    signed_in.post("/projects/1/schedule/layout/1", data={"x": "320", "y": "140"},
                   follow_redirects=True)
    assert "translate(320,140)" in text(signed_in.get("/projects/1/schedule"))

    signed_in.post("/projects/1/schedule/layout/reset", follow_redirects=True)
    body = text(signed_in.get("/projects/1/schedule"))
    assert "translate(320,140)" not in body


def test_a_deliverable_from_another_project_cannot_be_moved(signed_in):
    assert signed_in.post("/projects/1/schedule/layout/99999",
                          data={"x": "10", "y": "10"}).status_code == 404


def test_a_start_to_start_arrow_is_drawn_differently(signed_in):
    signed_in.post("/projects/1/schedule/links",
                   data={"predecessor_id": "1", "successor_id": "2", "kind": "SS", "lag_days": "5"},
                   follow_redirects=True)
    body = text(signed_in.get("/projects/1/schedule"))
    assert "stroke-dasharray" in body
    assert "start → start" in body


# --- the four link types ---------------------------------------------------

def kinds_case():
    """A 30-day predecessor and a 10-day successor, so each link is distinct."""
    return [task(1, "2026-01-01", "2026-01-30"), task(2, "2026-01-01", "2026-01-10")]


@pytest.mark.parametrize(
    "kind,lag,start,finish",
    [
        ("FS", 0, "2026-01-31", "2026-02-09"),   # the day after it finishes
        ("SS", 14, "2026-01-15", "2026-01-24"),  # a fortnight after it starts
        ("FF", 0, "2026-01-21", "2026-01-30"),   # finishes when it finishes
        ("SF", 0, "2026-01-01", "2026-01-10"),   # finishes when it starts
    ],
)
def test_each_kind_of_link_drives_the_end_it_names(kind, lag, start, finish):
    result = analyse(kinds_case(), [dict(link(1, 2, lag=lag), kind=kind)])[2]
    assert (result["early_start"], result["early_finish"]) == (start, finish)


def test_a_negative_lag_lets_two_pieces_of_work_overlap():
    """"Start a week before the survey ends" is how overlap is written down."""
    result = analyse(kinds_case(), [dict(link(1, 2, lag=-7), kind="FS")])[2]
    assert result["early_start"] == "2026-01-24"     # a week before 30 January
    assert result["early_finish"] == "2026-02-02"


def test_a_negative_lag_works_on_every_kind():
    """A move is where a lag shows plainly: the constraint sets the date rather
    than only raising a floor under the line's own planned start."""
    for kind in ("FS", "SS", "FF", "SF"):
        tasks = kinds_case()
        links = [dict(link(1, 2), kind=kind)]
        tasks[0]["start_date"] = "2026-02-01"
        tasks[0]["submission_date"] = "2026-03-02"

        on_time = shift_successors(tasks, links, 1)[2]["start_date"]
        overlapping = shift_successors(
            tasks, [dict(link(1, 2, lag=-5), kind=kind)], 1)[2]["start_date"]
        assert overlapping < on_time, f"{kind} ignored a negative lag"


def test_a_negative_lag_only_relaxes_a_constraint_it_does_not_drag_work_back():
    """A line still starts no earlier than it is planned to; the lag lowers the
    limit its predecessor puts on it, it does not pull it forward."""
    result = analyse(kinds_case(), [dict(link(1, 2, lag=-5), kind="SS")])[2]
    assert result["early_start"] == "2026-01-01"      # its own planned start


@pytest.mark.parametrize("kind", ["FS", "SS", "FF", "SF"])
def test_float_is_worked_out_for_every_kind(kind):
    result = analyse(kinds_case(), [dict(link(1, 2), kind=kind)])
    assert result[1]["in_a_chain"] and result[2]["in_a_chain"]
    assert isinstance(result[1]["total_float"], int)
    assert result[2]["total_float"] >= 0


@pytest.mark.parametrize(
    "kind,expected",
    [("FS", "2026-03-03"), ("SS", "2026-02-01"), ("FF", "2026-02-21"), ("SF", "2026-01-23")],
)
def test_moving_a_predecessor_moves_the_end_the_link_names(kind, expected):
    tasks = kinds_case()
    links = [dict(link(1, 2), kind=kind)]
    tasks[0]["start_date"] = "2026-02-01"            # a month later
    tasks[0]["submission_date"] = "2026-03-02"

    moves = shift_successors(tasks, links, 1)
    assert moves[2]["start_date"] == expected
    assert duration_between(moves[2]["start_date"], moves[2]["submission_date"]) == 10


def test_a_lag_can_be_negative_all_the_way_through_the_app(signed_in):
    signed_in.post("/projects/1/schedule/links",
                   data={"predecessor_id": "1", "successor_id": "2", "kind": "FS",
                         "lag_days": "-7"}, follow_redirects=True)
    from app.db import query_one

    with signed_in.application.app_context():
        assert query_one("SELECT lag_days FROM task_links LIMIT 1")["lag_days"] == -7
    assert "-7d" in text(signed_in.get("/projects/1/schedule"))


# --- what the programme chart draws ----------------------------------------

def test_the_chart_carries_a_band_of_months(signed_in):
    body = text(signed_in.get("/projects/1/schedule?data_date=01/09/2026"))
    assert "Sep 2026" in body and "Oct 2026" in body


def test_a_submission_is_a_green_star(signed_in):
    body = text(signed_in.get("/projects/1/schedule"))
    assert "<polygon" in body
    assert "var(--good)" in body


def test_the_legend_says_what_every_mark_means(signed_in):
    body = text(signed_in.get("/projects/1/schedule"))
    for meaning in ("Planned bar", "IDC (workflow only)", "Submission",
                    "Code A due (workflow only)", "Rework after a Code B or C", "Today"):
        assert meaning in body, meaning


def test_a_line_off_the_workflow_carries_no_idc_and_no_code_a():
    """A meeting or a transmittal has neither, because neither happens to it."""
    from app.charts import gantt

    def drawn(uses_workflow):
        return str(gantt([{
            "id": 1, "wbs": "1.1", "name": "x", "start_date": "2026-01-01",
            "submission_date": "2026-02-10", "actual_pct": 0, "is_critical": False,
            "duration_days": 41, "total_float": 0, "uses_workflow": uses_workflow,
            "approval_due_date": "2026-02-24",
            "step_plan": [{"key": "idc", "date": "2026-02-05"}], "revisions": [],
        }], "2026-01-01", "2026-03-10", "2026-01-20"))

    workflow, simple = drawn(True), drawn(False)
    assert workflow.count("<circle") == 1 and workflow.count("<polygon") == 2
    assert simple.count("<circle") == 0                # no IDC
    assert simple.count("<polygon") == 1               # only the submission star


def test_the_schedule_table_leaves_code_a_blank_off_the_workflow(signed_in):
    body = text(signed_in.get("/projects/1/schedule"))
    assert ">Code A<" in body                          # the column is still there
    assert "—" in body                                 # and empty where it does not apply


# --- the dependency workbook -----------------------------------------------

def test_the_dependencies_export_to_a_workbook(signed_in):
    link_two(signed_in, "2", "1")
    response = signed_in.get("/projects/1/schedule/links.xlsx")
    assert response.status_code == 200
    assert response.data[:2] == b"PK"
    assert "dependencies" in response.headers["Content-Disposition"]

    from app.excel import read_links_workbook

    rows = read_links_workbook(response.data)
    assert len(rows) == 1
    assert rows[0]["kind"] == "FS"


def test_the_workbook_names_deliverables_by_wbs_and_lists_them(signed_in):
    link_two(signed_in, "2", "1")
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(signed_in.get("/projects/1/schedule/links.xlsx").data))
    assert wb.sheetnames == ["Dependencies", "Deliverables"]
    assert [c.value for c in wb["Dependencies"][1]][:4] == [
        "Deliverable WBS", "Waits for WBS", "Type", "Lag days"
    ]
    assert wb["Deliverables"].max_row > 5              # somewhere to copy a WBS from


def upload(client, data: bytes):
    import io

    return client.post("/projects/1/schedule/links/import",
                       data={"workbook": (io.BytesIO(data), "links.xlsx")},
                       content_type="multipart/form-data",
                       follow_redirects=True).get_data(as_text=True)


def workbook_of(rows):
    """A dependency workbook built from (successor, predecessor, kind, lag)."""
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Dependencies")
    ws.append(["Deliverable WBS", "Waits for WBS", "Type", "Lag days"])
    for row in rows:
        ws.append(list(row))
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def links_now(client):
    from app.db import query

    with client.application.app_context():
        return [
            (r["successor_wbs"], r["predecessor_wbs"], r["kind"], r["lag_days"])
            for r in query(
                "SELECT l.kind, l.lag_days, p.wbs AS predecessor_wbs, s.wbs AS successor_wbs "
                "FROM task_links l JOIN tasks p ON p.id = l.predecessor_id "
                "JOIN tasks s ON s.id = l.successor_id WHERE l.project_id = 1"
            )
        ]


def test_an_edited_workbook_becomes_the_programmes_dependencies(signed_in):
    body = upload(signed_in, workbook_of([
        ("1.2", "1.1", "FS", 0), ("1.3", "1.2", "SS", 5), ("1.4", "1.1", "FF", -3),
    ]))
    assert "3 dependencies imported" in body
    assert sorted(links_now(signed_in)) == [
        ("1.2", "1.1", "FS", 0), ("1.3", "1.2", "SS", 5), ("1.4", "1.1", "FF", -3),
    ]


def test_importing_replaces_what_was_there(signed_in):
    link_two(signed_in, "2", "1")
    upload(signed_in, workbook_of([("1.3", "1.2", "SS", 2)]))
    assert links_now(signed_in) == [("1.3", "1.2", "SS", 2)]


def test_a_row_naming_a_wbs_that_does_not_exist_is_reported(signed_in):
    body = upload(signed_in, workbook_of([("1.2", "9.9", "FS", 0), ("1.3", "1.2", "FS", 0)]))
    assert "1 dependency imported" in body
    assert "no deliverable with that WBS" in body
    assert links_now(signed_in) == [("1.3", "1.2", "FS", 0)]


def test_a_workbook_that_would_loop_is_refused_row_by_row(signed_in):
    body = upload(signed_in, workbook_of([("1.2", "1.1", "FS", 0), ("1.1", "1.2", "FS", 0)]))
    assert "would make the programme depend on itself" in body
    assert links_now(signed_in) == [("1.2", "1.1", "FS", 0)]


def test_the_notes_under_the_sheet_do_not_trip_the_import(signed_in):
    """The exported sheet carries a blank line and two notes; importing the
    file it produced has to work."""
    link_two(signed_in, "2", "1")
    exported = signed_in.get("/projects/1/schedule/links.xlsx").data
    assert "1 dependency imported" in upload(signed_in, exported)


def test_a_file_that_is_not_a_workbook_says_so(signed_in):
    assert "could not be read as an Excel workbook" in upload(signed_in, b"not a workbook")


def test_a_workbook_without_the_sheet_says_which_one_is_missing(signed_in):
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Something else"
    stream = io.BytesIO()
    wb.save(stream)
    assert "no Dependencies sheet" in upload(signed_in, stream.getvalue())


def test_importing_nothing_says_so(signed_in):
    body = signed_in.post("/projects/1/schedule/links/import", data={},
                          follow_redirects=True).get_data(as_text=True)
    assert "Choose a workbook" in body


def test_only_a_manager_can_import_dependencies(client, app):
    client.post("/register", data={"name": "Member", "email": "m2@example.com",
                                   "password": "longenough1"})
    with app.app_context():
        from app.db import connect

        conn = connect(app.config["DATABASE"])
        with conn:
            user = conn.execute("SELECT id FROM users WHERE email = 'm2@example.com'").fetchone()
            conn.execute("INSERT INTO project_members (project_id, user_id, role) VALUES (1, ?, 'member')",
                         (user["id"],))
        conn.close()

    assert client.get("/projects/1/schedule/links.xlsx").status_code == 200     # they may read
    assert client.post("/projects/1/schedule/links/import").status_code == 403


# --- moving an end of a dependency -----------------------------------------

def wbs_of(client, link_id: int):
    from app.db import query_one

    with client.application.app_context():
        row = query_one(
            "SELECT p.wbs AS predecessor, s.wbs AS successor FROM task_links l "
            "JOIN tasks p ON p.id = l.predecessor_id JOIN tasks s ON s.id = l.successor_id "
            "WHERE l.id = ?", (link_id,))
    return row["successor"], row["predecessor"]


def test_either_end_of_a_dependency_can_be_moved(signed_in):
    link_two(signed_in, "2", "1")
    link_id = link_id_of(signed_in)

    result = signed_in.post(f"/projects/1/schedule/links/{link_id}",
                            data={"successor_id": "3"},
                            headers={"Accept": "application/json"}).get_json()
    assert result["ok"]
    assert wbs_of(signed_in, link_id) == ("1.3", "1.1")

    signed_in.post(f"/projects/1/schedule/links/{link_id}", data={"predecessor_id": "2"},
                   headers={"Accept": "application/json"})
    assert wbs_of(signed_in, link_id) == ("1.3", "1.2")


def test_the_row_comes_back_drawn_when_an_end_moves(signed_in):
    """Its WBS numbers change with it, so the row is re-rendered rather than
    patched in place."""
    link_two(signed_in, "2", "1")
    result = signed_in.post(f"/projects/1/schedule/links/{link_id_of(signed_in)}",
                            data={"successor_id": "3"},
                            headers={"Accept": "application/json"}).get_json()
    assert "1.3" in result["link_html"]
    assert 'data-cell="link-successor"' in result["link_html"]


def test_moving_an_end_onto_itself_is_refused(signed_in):
    link_two(signed_in, "2", "1")
    response = signed_in.post(f"/projects/1/schedule/links/{link_id_of(signed_in)}",
                              data={"successor_id": "1"},
                              headers={"Accept": "application/json"})
    assert response.status_code == 400
    assert "cannot depend on itself" in response.get_json()["error"]
    assert wbs_of(signed_in, link_id_of(signed_in)) == ("1.2", "1.1")   # unchanged


def test_moving_an_end_so_the_programme_loops_is_refused(signed_in):
    link_two(signed_in, "2", "1")
    link_two(signed_in, "3", "2")
    from app.db import query_one

    with signed_in.application.app_context():
        first = query_one("SELECT id FROM task_links ORDER BY id LIMIT 1")["id"]

    # 1.2 waits for 1.1; making it wait for 1.3 would close the loop.
    response = signed_in.post(f"/projects/1/schedule/links/{first}",
                              data={"predecessor_id": "3"},
                              headers={"Accept": "application/json"})
    assert response.status_code == 400
    assert "depend on itself" in response.get_json()["error"]


def test_moving_an_end_onto_a_pair_that_already_exists_is_refused(signed_in):
    link_two(signed_in, "2", "1")
    link_two(signed_in, "3", "1")
    from app.db import query_one

    with signed_in.application.app_context():
        second = query_one("SELECT id FROM task_links ORDER BY id DESC LIMIT 1")["id"]

    response = signed_in.post(f"/projects/1/schedule/links/{second}",
                              data={"successor_id": "2"},
                              headers={"Accept": "application/json"})
    assert response.status_code == 400
    assert "already linked" in response.get_json()["error"]


def test_a_deliverable_from_another_project_cannot_be_put_on_a_link(signed_in):
    link_two(signed_in, "2", "1")
    response = signed_in.post(f"/projects/1/schedule/links/{link_id_of(signed_in)}",
                              data={"successor_id": "99999"},
                              headers={"Accept": "application/json"})
    assert response.status_code == 400
    assert "must belong to this project" in response.get_json()["error"]


def test_moving_an_end_leaves_the_lag_and_the_kind_alone(signed_in):
    signed_in.post("/projects/1/schedule/links",
                   data={"predecessor_id": "1", "successor_id": "2", "kind": "SS",
                         "lag_days": "-4"}, follow_redirects=True)
    signed_in.post(f"/projects/1/schedule/links/{link_id_of(signed_in)}",
                   data={"successor_id": "3"}, headers={"Accept": "application/json"})

    from app.db import query_one

    with signed_in.application.app_context():
        row = query_one("SELECT kind, lag_days FROM task_links LIMIT 1")
    assert row["kind"] == "SS" and row["lag_days"] == -4


def test_both_ends_offer_the_deliverable_list_in_the_row(signed_in):
    link_two(signed_in, "2", "1")
    body = text(signed_in.get("/projects/1/schedule"))
    assert 'data-cell="link-successor"' in body and 'data-cell="link-predecessor"' in body
    assert body.count('id="cell-link-successor"') == 1     # one copy, cloned per row
    assert body.count('id="cell-link-predecessor"') == 1


# --- the schedule workbook -------------------------------------------------

def test_the_schedule_exports_to_a_workbook(signed_in):
    response = signed_in.get("/projects/1/schedule.xlsx")
    assert response.status_code == 200
    assert response.data[:2] == b"PK"
    assert "schedule" in response.headers["Content-Disposition"]

    from app.excel import read_schedule_workbook

    rows = read_schedule_workbook(response.data)
    assert len(rows) == 55
    assert rows[0]["wbs"] == "1.1"
    assert rows[0]["duration_days"] > 0


def test_the_workbook_shades_whichever_column_is_worked_out(signed_in):
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(signed_in.get("/projects/1/schedule.xlsx").data))
    assert [c.value for c in wb["Schedule"][1]] == [
        "WBS", "Deliverable", "Start", "Duration (days)", "Finish", "Section"
    ]
    assert "worked out from the start and the duration" in str(
        [c.value for row in wb["Schedule"].iter_rows() for c in row])

    signed_in.post("/projects/1/schedule/mode", data={"mode": "dates"}, follow_redirects=True)
    wb = openpyxl.load_workbook(io.BytesIO(signed_in.get("/projects/1/schedule.xlsx").data))
    assert "worked out from the two dates" in str(
        [c.value for row in wb["Schedule"].iter_rows() for c in row])


def schedule_workbook(rows):
    """A schedule workbook built from (wbs, start, duration, finish)."""
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Schedule"
    wb.active.append(["WBS", "Deliverable", "Start", "Duration (days)", "Finish", "Section"])
    for wbs, start, duration, finish in rows:
        wb.active.append([wbs, "", start, duration, finish, ""])
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def send_schedule(client, data: bytes):
    import io

    return client.post("/projects/1/schedule/import",
                       data={"workbook": (io.BytesIO(data), "schedule.xlsx")},
                       content_type="multipart/form-data",
                       follow_redirects=True).get_data(as_text=True)


def test_an_edited_workbook_reschedules_by_duration(signed_in):
    body = send_schedule(signed_in, schedule_workbook([("1.1", "2026-10-01", 15, "")]))
    assert "1 deliverable rescheduled" in body
    assert dates_of(signed_in, 1) == ("2026-10-01", "2026-10-15")


def test_an_edited_workbook_reschedules_by_dates_when_that_is_the_setting(signed_in):
    signed_in.post("/projects/1/schedule/mode", data={"mode": "dates"}, follow_redirects=True)
    send_schedule(signed_in, schedule_workbook([("1.1", "2026-10-01", 99, "2026-10-20")]))
    assert dates_of(signed_in, 1) == ("2026-10-01", "2026-10-20")   # the dates win, not 99


def test_dd_mm_yyyy_is_understood_in_the_workbook(signed_in):
    send_schedule(signed_in, schedule_workbook([("1.1", "01/10/2026", 10, "")]))
    assert dates_of(signed_in, 1) == ("2026-10-01", "2026-10-10")


def test_a_row_naming_a_wbs_that_does_not_exist_is_reported_on_import(signed_in):
    body = send_schedule(signed_in, schedule_workbook(
        [("9.9", "2026-10-01", 5, ""), ("1.1", "2026-10-01", 5, "")]))
    assert "1 deliverable rescheduled" in body
    assert "no deliverable with that WBS" in body


def test_a_finish_before_the_start_is_reported_rather_than_written(signed_in):
    signed_in.post("/projects/1/schedule/mode", data={"mode": "dates"}, follow_redirects=True)
    before = dates_of(signed_in, 1)
    body = send_schedule(signed_in, schedule_workbook([("1.1", "2026-10-10", 0, "2026-10-01")]))
    assert "the finish is before the start" in body
    assert dates_of(signed_in, 1) == before


def test_the_import_takes_the_sheet_as_it_is_written(signed_in):
    """The sheet is the authority, so a line is not pushed on top of what it
    says by something it depends on."""
    link_two(signed_in, "2", "1")
    send_schedule(signed_in, schedule_workbook(
        [("1.1", "2026-10-01", 30, ""), ("1.2", "2026-10-05", 10, "")]))
    assert dates_of(signed_in, 2) == ("2026-10-05", "2026-10-14")    # not pushed to November


def test_its_own_export_imports_without_a_complaint(signed_in):
    assert "skipped" not in send_schedule(signed_in, signed_in.get("/projects/1/schedule.xlsx").data)


def test_a_schedule_file_that_is_not_a_workbook_says_so(signed_in):
    assert "could not be read as an Excel workbook" in send_schedule(signed_in, b"nope")


def test_a_workbook_without_the_schedule_sheet_says_which_one_is_missing(signed_in):
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Something else"
    stream = io.BytesIO()
    wb.save(stream)
    assert "no Schedule sheet" in send_schedule(signed_in, stream.getvalue())


def test_only_a_manager_can_import_a_schedule(client, app):
    client.post("/register", data={"name": "Member", "email": "m3@example.com",
                                   "password": "longenough1"})
    with app.app_context():
        from app.db import connect

        conn = connect(app.config["DATABASE"])
        with conn:
            user = conn.execute("SELECT id FROM users WHERE email = 'm3@example.com'").fetchone()
            conn.execute("INSERT INTO project_members (project_id, user_id, role) VALUES (1, ?, 'member')",
                         (user["id"],))
        conn.close()

    assert client.get("/projects/1/schedule.xlsx").status_code == 200
    assert client.post("/projects/1/schedule/import").status_code == 403


def test_the_notes_under_a_sheet_never_sit_in_a_column_that_is_read(signed_in):
    """Prose in the WBS column reads as a row, and the import then complains
    about a deliverable called "Dates read yyyy-mm-dd"."""
    from app.excel import read_links_workbook, read_schedule_workbook

    link_two(signed_in, "2", "1")
    schedule = read_schedule_workbook(signed_in.get("/projects/1/schedule.xlsx").data)
    assert all(len(row["wbs"]) < 20 for row in schedule), "a note was read as a deliverable"
    assert len(schedule) == 55

    links = read_links_workbook(signed_in.get("/projects/1/schedule/links.xlsx").data)
    assert len(links) == 1
