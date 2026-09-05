"""Exporting the whole project setup to a workbook, and reading it back.

The workbook is the editing surface people already know: change weights, dates,
the trade split or the workflow in Excel, then import it. Export and import use
the same sheet layout, so a file that came out can always go back in.

openpyxl is imported lazily so the rest of the app still runs if it is missing.
"""

from __future__ import annotations

import io
from typing import Any, Mapping, Sequence

from .dates import from_input, to_display

SHEET_PROJECT = "Project"
SHEET_WORKFLOW = "Workflow"
SHEET_TRADES = "Trades"
SHEET_SECTIONS = "Sections"
SHEET_TASKS = "Deliverables"

HELP = (
    "Blue headings are read on import. Grey columns are for reference only and are "
    "ignored — including Status and Revision, so importing never overwrites progress "
    "reported in the app."
)


class ExcelUnavailable(RuntimeError):
    """openpyxl is not installed."""


class ImportError_(ValueError):
    """The uploaded workbook could not be used."""


def _openpyxl():
    try:
        import openpyxl  # noqa: PLC0415 - optional dependency, loaded on demand
    except ModuleNotFoundError as exc:  # pragma: no cover - depends on the install
        raise ExcelUnavailable(
            "Excel export and import need the openpyxl package. Install it with:\n"
            "    pip install openpyxl"
        ) from exc
    return openpyxl


# --- export ----------------------------------------------------------------

def build_workbook(
    project: Mapping[str, Any],
    steps: Sequence[Mapping[str, Any]],
    trades: Sequence[Mapping[str, Any]],
    sections: Sequence[Mapping[str, Any]],
    tasks: Sequence[Mapping[str, Any]],
) -> bytes:
    """The whole setup as an .xlsx file."""
    openpyxl = _openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A78D6")
    calc_fill = PatternFill("solid", fgColor="E1E0D9")
    note_font = Font(italic=True, color="898781")

    def sheet(title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]],
              calculated: Sequence[int] = (), widths: Sequence[int] = ()) -> None:
        ws = wb.create_sheet(title)
        ws.append(list(headers))
        for index, cell in enumerate(ws[1]):
            cell.font = header_font
            cell.fill = calc_fill if index in calculated else header_fill
            if index in calculated:
                cell.font = Font(bold=True, color="52514E")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append(list(row))
        ws.freeze_panes = "A2"
        for index, width in enumerate(widths or [18] * len(headers)):
            ws.column_dimensions[get_column_letter(index + 1)].width = width

    # Project settings, as a name/value list so it stays readable.
    ws = wb.active
    ws.title = SHEET_PROJECT
    ws.append(["Setting", "Value", "Notes"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    settings = [
        ("Project code", project["code"], "Must stay unique"),
        ("Project name", project["name"], ""),
        ("Client", project["client"], ""),
        ("Description", project["description"], ""),
        ("Notice to proceed", to_display(project["ntp_date"]), "dd/mm/yyyy"),
        ("Duration (months)", project["duration_months"], ""),
        ("Days per month", project["days_per_month"], "Programme basis for day counts"),
        ("Hours per man-month", project["hours_per_month"], "Converts hour budgets to man-months"),
        ("Count NTP day as elapsed", int(project["elapsed_day_offset"]), "0 or 1"),
        ("Maximum revisions", project["max_revisions"], "Resubmissions before escalation"),
        ("Rework days", project["rework_days"], "Comments received to next submission"),
        ("Revision resets to", project["revision_reset_step"], "Workflow step key"),
        ("Status", project["status"], "active / on_hold / complete / archived"),
    ]
    for row in settings:
        ws.append(list(row))
    ws.append([])
    ws.append([HELP])
    ws.cell(row=ws.max_row, column=1).font = note_font
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 40
    ws.freeze_panes = "A2"

    sheet(
        SHEET_WORKFLOW,
        ["Step", "Percent complete", "Anchor", "Offset (days)"],
        [[s["name"], round(float(s["percent"]) * 100, 4), s["anchor"], s["offset_days"]] for s in steps],
        widths=[28, 18, 14, 16],
    )
    ws = wb[SHEET_WORKFLOW]
    ws.append([])
    ws.append(["Anchor is 'start' or 'submission'. A negative offset is before that date."])
    ws.cell(row=ws.max_row, column=1).font = note_font

    sheet(
        SHEET_TRADES,
        ["Trade", "Budget (hours)", "Colour"],
        [[t["name"], t["budget_hours"], t["color"]] for t in trades],
        widths=[28, 16, 12],
    )
    sheet(
        SHEET_SECTIONS,
        ["Code", "Section"],
        [[s["code"], s["name"]] for s in sections],
        widths=[12, 52],
    )

    section_name = {s["id"]: s["name"] for s in sections}
    trade_names = [t["name"] for t in trades]
    headers = (
        ["WBS", "Section", "Deliverable", "Weight points", "Start date", "Submission date",
         "Tracking", "Status", "Revision", "Remarks"]
        + [f"{name} %" for name in trade_names]
        + ["Split total %", "Weight %"]
    )
    # Status and Revision are shown for reference but are never read back: progress
    # is reported in the app, and a workbook edited offline would otherwise revert
    # anything reported since it was exported.
    calculated = {7, 8, len(headers) - 2, len(headers) - 1}

    rows = []
    total_points = sum(float(t["weight_points"] or 0) for t in tasks) or 1.0
    for task in tasks:
        allocations = task.get("allocations") or {}
        shares = [round(float(allocations.get(t["id"], 0)) * 100, 4) for t in trades]
        rows.append(
            [task["wbs"], section_name.get(task["section_id"], ""), task["name"],
             task["weight_points"], to_display(task["start_date"]), to_display(task["submission_date"]),
             task["tracking"], task["status_key"], task["revision"], task["remarks"]]
            + shares
            + [round(sum(shares), 4), round(float(task["weight_points"] or 0) / total_points * 100, 4)]
        )
    sheet(
        SHEET_TASKS, headers, rows, calculated=calculated,
        widths=[10, 26, 60, 14, 14, 16, 12, 20, 10, 30] + [14] * len(trade_names) + [14, 12],
    )

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


# --- import ----------------------------------------------------------------

def _cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any, default: float = 0.0) -> float:
    if value is None or str(value).strip() == "":
        return default
    try:
        return float(str(value).strip().rstrip("%"))
    except ValueError:
        return default


def read_workbook(data: bytes) -> dict[str, Any]:
    """Parses an exported workbook back into plain data.

    Nothing is written here — the caller applies the result, so a malformed file
    cannot leave a project half-updated.
    """
    openpyxl = _openpyxl()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises a variety of errors
        raise ImportError_("That file could not be read as an Excel workbook.") from exc

    missing = [name for name in (SHEET_TASKS, SHEET_TRADES) if name not in wb.sheetnames]
    if missing:
        raise ImportError_(
            f"The workbook is missing the {' and '.join(missing)} sheet. "
            "Export the setup first and edit that file."
        )

    out: dict[str, Any] = {"project": {}, "steps": [], "trades": [], "sections": [], "tasks": []}

    if SHEET_PROJECT in wb.sheetnames:
        labels = {
            "project code": "code", "project name": "name", "client": "client",
            "description": "description", "notice to proceed": "ntp_date",
            "duration (months)": "duration_months", "days per month": "days_per_month",
            "hours per man-month": "hours_per_month", "count ntp day as elapsed": "elapsed_day_offset",
            "maximum revisions": "max_revisions", "rework days": "rework_days",
            "revision resets to": "revision_reset_step", "status": "status",
        }
        for row in wb[SHEET_PROJECT].iter_rows(min_row=2, values_only=True):
            key = labels.get(_cell(row[0]).lower())
            if key and len(row) > 1:
                out["project"][key] = row[1]

    if SHEET_WORKFLOW in wb.sheetnames:
        for row in wb[SHEET_WORKFLOW].iter_rows(min_row=2, values_only=True):
            name = _cell(row[0])
            if not name or name.lower().startswith("anchor is"):
                continue
            anchor = _cell(row[2]).lower() or "submission"
            out["steps"].append(
                {
                    "name": name,
                    "percent": max(0.0, min(1.0, _number(row[1]) / 100)),
                    "anchor": "start" if anchor.startswith("start") else "submission",
                    "offset_days": _number(row[3]),
                }
            )

    for row in wb[SHEET_TRADES].iter_rows(min_row=2, values_only=True):
        name = _cell(row[0])
        if name:
            out["trades"].append(
                {"name": name, "budget_hours": _number(row[1]), "color": _cell(row[2]) or "#2a78d6"}
            )

    if SHEET_SECTIONS in wb.sheetnames:
        for row in wb[SHEET_SECTIONS].iter_rows(min_row=2, values_only=True):
            name = _cell(row[1]) if len(row) > 1 else ""
            if name:
                out["sections"].append({"code": _cell(row[0]), "name": name})

    ws = wb[SHEET_TASKS]
    header = [_cell(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    trade_columns = {}
    for index, title in enumerate(header):
        if title.endswith(" %") and title[:-2] not in ("Split total", "Weight"):
            trade_columns[title[:-2]] = index

    for number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = _cell(row[2]) if len(row) > 2 else ""
        if not name:
            continue
        start = from_input(row[4]) if len(row) > 4 else None
        submission = from_input(row[5]) if len(row) > 5 else None
        if not submission and not start:
            raise ImportError_(f"Row {number} (“{name[:40]}”) has no start or submission date.")

        allocations = {
            trade: round(_number(row[index]) / 100, 6)
            for trade, index in trade_columns.items()
            if index < len(row) and _number(row[index]) > 0
        }
        total = sum(allocations.values())
        if allocations and abs(total - 1) > 0.005:
            raise ImportError_(
                f"Row {number} (“{name[:40]}”): the trade split totals {total * 100:.1f}%, not 100%."
            )

        out["tasks"].append(
            {
                "wbs": _cell(row[0]),
                "section": _cell(row[1]),
                "name": name,
                "weight_points": _number(row[3]),
                "start_date": start or submission,
                "submission_date": submission or start,
                "tracking": (_cell(row[6]) or "workflow").lower(),
                # Status and Revision are exported for reference only.
                "remarks": _cell(row[9]) if len(row) > 9 else "",
                "allocations": allocations,
            }
        )

    if not out["tasks"]:
        raise ImportError_("The Deliverables sheet has no rows with a name.")
    return out


# --- the dependency workbook -----------------------------------------------

SHEET_LINKS = "Dependencies"


def build_links_workbook(project: Mapping[str, Any], tasks: Sequence[Mapping[str, Any]],
                         links: Sequence[Mapping[str, Any]]) -> bytes:
    """The programme's dependencies as an .xlsx file, ready to edit and import.

    Deliverables are named by WBS, which is what a reader recognises and what
    the import matches on. A second sheet lists the WBS numbers to copy from, so
    nobody has to guess at one.
    """
    openpyxl = _openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]],
              widths: Sequence[int]) -> None:
        ws = wb.create_sheet(title)
        ws.append(list(headers))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2A78D6")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for row in rows:
            ws.append(list(row))
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + index)].width = width
        ws.freeze_panes = "A2"

    from .schedule import KIND_KEYS, kind_note, normalise_kind

    sheet(
        SHEET_LINKS,
        ["Deliverable WBS", "Waits for WBS", "Type", "Lag days", "What it means"],
        [
            [link.get("successor_wbs"), link.get("predecessor_wbs"),
             normalise_kind(link.get("kind")), float(link.get("lag_days") or 0),
             kind_note(link.get("kind"))]
            for link in links
        ],
        widths=[18, 18, 10, 12, 44],
    )

    guide = wb[SHEET_LINKS]
    guide.append([])
    guide.append(["", "Type is one of " + ", ".join(KIND_KEYS)
                      + ". Lag may be negative, for work that overlaps."])
    guide.append(["", "Importing replaces every dependency with what is in this sheet."])
    for row in guide.iter_rows(min_row=guide.max_row - 1, max_row=guide.max_row):
        row[1].font = Font(italic=True, color="898781")

    sheet(
        "Deliverables",
        ["WBS", "Deliverable", "Start", "Finish"],
        [[t.get("wbs"), t.get("name"), _cell(t.get("start_date")), _cell(t.get("submission_date"))]
         for t in tasks],
        widths=[14, 62, 14, 14],
    )
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def read_links_workbook(data: bytes) -> list[dict[str, Any]]:
    """The dependency sheet as plain rows. Nothing is written here.

    Rows without both WBS numbers are skipped, which is what lets the sheet
    carry the blank line and the notes underneath without tripping the import.
    """
    from .schedule import normalise_kind

    openpyxl = _openpyxl()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises a variety of errors
        raise ImportError_("That file could not be read as an Excel workbook.") from exc

    if SHEET_LINKS not in wb.sheetnames:
        raise ImportError_(
            f"The workbook has no {SHEET_LINKS} sheet. Export the dependencies "
            "first and edit that file."
        )

    rows: list[dict[str, Any]] = []
    for row in wb[SHEET_LINKS].iter_rows(min_row=2, values_only=True):
        successor = _cell(row[0] if len(row) > 0 else "")
        predecessor = _cell(row[1] if len(row) > 1 else "")
        if not successor or not predecessor:
            continue
        rows.append({
            "successor_wbs": successor,
            "predecessor_wbs": predecessor,
            "kind": normalise_kind(row[2] if len(row) > 2 else ""),
            "lag_days": _number(row[3] if len(row) > 3 else 0, 0.0),
        })
    return rows


# --- the schedule workbook -------------------------------------------------

SHEET_SCHEDULE = "Schedule"


def build_schedule_workbook(project: Mapping[str, Any], tasks: Sequence[Mapping[str, Any]],
                            mode: str = "duration") -> bytes:
    """The programme's dates as an .xlsx file, ready to edit and import back.

    Which two of start, duration and finish you edit follows the project's own
    setting, exactly as on screen: by duration the finish is worked out, by
    dates the duration is. The calculated one is shaded, so there is no doubt
    about which column is read on the way back in.
    """
    openpyxl = _openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_SCHEDULE

    by_duration = mode != "dates"
    calculated = 4 if by_duration else 3          # Finish, or Duration
    headers = ["WBS", "Deliverable", "Start", "Duration (days)", "Finish", "Section"]

    ws.append(headers)
    for index, cell in enumerate(ws[1]):
        if index in (calculated, 5):
            cell.font = Font(bold=True, color="52514E")
            cell.fill = PatternFill("solid", fgColor="E1E0D9")
        else:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2A78D6")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    from .schedule import duration_between

    for task in tasks:
        ws.append([
            _cell(task.get("wbs")),
            _cell(task.get("name")),
            _cell(task.get("start_date")),
            duration_between(task.get("start_date"), task.get("submission_date")),
            _cell(task.get("submission_date")),
            _cell(task.get("section_name")),
        ])

    for column, width in zip("ABCDEF", (14, 62, 14, 16, 14, 30)):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"

    ws.append([])
    note = ("Dates read yyyy-mm-dd or dd/mm/yyyy. "
            + ("Finish is worked out from the start and the duration — edit those two."
               if by_duration else
               "Duration is worked out from the two dates — edit the start and the finish."))
    ws.append(["", note])
    ws.append(["", "Deliverables are matched on WBS. A row whose WBS is not on the project is "
                   "reported rather than added."])
    for row in ws.iter_rows(min_row=ws.max_row - 1, max_row=ws.max_row):
        row[1].font = Font(italic=True, color="898781")

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def read_schedule_workbook(data: bytes) -> list[dict[str, Any]]:
    """The schedule sheet as plain rows. Nothing is written here.

    Rows without a WBS are skipped, which is what lets the notes under the
    table ride along in the same file.
    """
    openpyxl = _openpyxl()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises a variety of errors
        raise ImportError_("That file could not be read as an Excel workbook.") from exc

    if SHEET_SCHEDULE not in wb.sheetnames:
        raise ImportError_(
            f"The workbook has no {SHEET_SCHEDULE} sheet. Export the schedule "
            "first and edit that file."
        )

    rows: list[dict[str, Any]] = []
    for row in wb[SHEET_SCHEDULE].iter_rows(min_row=2, values_only=True):
        wbs = _cell(row[0] if len(row) > 0 else "")
        if not wbs:
            continue
        rows.append({
            "wbs": wbs,
            "start_date": _cell(row[2] if len(row) > 2 else ""),
            "duration_days": _number(row[3] if len(row) > 3 else 0, 0.0),
            "submission_date": _cell(row[4] if len(row) > 4 else ""),
        })
    return rows
