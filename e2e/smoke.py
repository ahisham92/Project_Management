"""End-to-end smoke test: signs in, reads the dashboard, records progress, books
hours, and checks both themes and the mobile layout.

    pip install playwright && python -m playwright install chromium
    python run.py seed && python run.py          # in one terminal
    python e2e/smoke.py                          # in another

Run it against a freshly seeded database — it books hours, so repeated runs
against the same database accumulate them and the budget check will fail.
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

from playwright.sync_api import sync_playwright

BASE = os.environ.get("BASE_URL", "http://localhost:8000")
EMAIL = os.environ.get("SEED_EMAIL", "admin@example.com")
PASSWORD = os.environ.get("SEED_PASSWORD", "changeme123")
SHOTS = Path(sys.argv[1] if len(sys.argv) > 1 else "e2e/screenshots")

failures: list[str] = []


def main() -> int:
    SHOTS.mkdir(parents=True, exist_ok=True)
    launch = {"executable_path": os.environ["CHROMIUM_PATH"]} if os.environ.get("CHROMIUM_PATH") else {}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(**launch)
        page = browser.new_page(viewport={"width": 1440, "height": 1000}, accept_downloads=True)
        page.on("console", lambda m: failures.append(f"console: {m.text}")
                if m.type == "error" and "400 (BAD REQUEST)" not in m.text else None)
        page.on("pageerror", lambda e: failures.append(f"pageerror: {e}"))

        def step(name: str, fn) -> None:
            try:
                fn(page)
                print(f"  PASS  {name}")
            except Exception as exc:  # noqa: BLE001 - the report is the point
                print(f"  FAIL  {name}: {exc}")
                failures.append(f"{name}: {exc}")

        def shot(name: str) -> None:
            page.screenshot(path=str(SHOTS / f"{name}.png"), full_page=True)

        step("login page renders", lambda p: (
            p.goto(f"{BASE}/login", wait_until="networkidle"),
            p.wait_for_selector("text=Project Control"),
            shot("01-login"),
        ))

        step("rejects a bad password", lambda p: (
            p.fill("input[name=email]", EMAIL),
            p.fill("input[name=password]", "totallywrong"),
            p.click("button[type=submit]"),
            p.wait_for_selector("text=Incorrect email or password", timeout=5000),
        ))

        step("signs in", lambda p: (
            p.fill("input[name=email]", EMAIL),
            p.fill("input[name=password]", PASSWORD),
            p.click("button[type=submit]"),
            p.wait_for_selector("text=Portfolio", timeout=8000),
            p.wait_for_selector("text=SIBLINE-PORT"),
            shot("02-portfolio"),
        ))

        step("portfolio shows the project and its status", lambda p: _expect_all(
            p, ["Sibline Port", "1 late", "Hours booked"]
        ))

        step("dashboard draws the S-curve with both series", _dashboard)
        step("trade table lists all four trades", lambda p: _expect_all(
            p, ["Marine", "Geotechnical", "Marine Structures", "Utilities"]
        ))
        step("chart tooltip appears on hover", _hover_tooltip)
        step("progress page lists deliverables", _progress_page)
        step("records a progress update by status, in the row", _record_progress)
        step("raises a revision when comments come back", _raise_revision)
        step("filters to late deliverables", _filter_late)
        step("schedule draws the programme with its milestones", _schedule)
        step("schedule links two deliverables and shifts what follows", _schedule_links)
        step("schedule dates and durations are amended in the row", _schedule_amend)
        step("schedule dependencies are edited and dragged", _schedule_deps)
        step("schedule dates go out to Excel and come back", _schedule_excel)
        step("schedule reads at a glance and folds its tables away", _schedule_reading)
        step("dependency lines are colour-coded, and Simplify untangles them", _schedule_simplify)
        step("dates read dd/mm/yyyy", _dates_read_dd_mm)
        step("budget page renders the hours chart", _budget)
        step("books hours and they reach budget control", _book_hours)
        step("period report shows what moved", _period)
        step("minutes: adds attendees, a meeting and its items", _minutes_capture)
        step("minutes: filters, searches and sorts the register", _minutes_filters)
        step("minutes: exports a set of minutes to Word", _minutes_word)
        step("minutes: reorders items and renumbers them", _minutes_reorder)
        step("minutes: edits an item in place without reloading", _minutes_edit_in_place)
        step("minutes: changes a field by clicking it in the row", _minutes_cells)
        step("minutes: picks a date from the calendar", _minutes_calendar)
        step("minutes: the agenda lists what is still open", _minutes_agenda)
        step("progress sorts by a column", _sorting)
        step("planned reads only the workflow step values", _stepped_planned)
        step("setup starts locked and opens with the password", _setup_lock)
        step("setup saves everything with one button", _save_all)
        step("setup exports to Excel", _excel_export)
        step("report tabs print to PDF", _print_to_pdf)
        step("dark mode renders", _dark)
        step("a change appears on another page without a refresh", _live)
        step("mobile layout does not overflow horizontally", _mobile)

        browser.close()

    print("\n" + ("ERRORS:" if failures else "No errors."))
    for failure in failures:
        print("  -", failure)
    return 1 if failures else 0


def _schedule_page(page, panel: str = ""):
    """The schedule, with the folded panel a step needs already open."""
    page.goto(f"{BASE}/projects/1/schedule" + (f"?panel={panel}" if panel else ""),
              wait_until="networkidle")


def _card(page, heading: str):
    """The card under a heading — several of them carry an Export button now."""
    return page.locator("div.card").filter(has=page.locator(f"h2:text-is('{heading}')"))


def _expect_all(page, needles: list[str]) -> None:
    body = page.text_content("body")
    missing = [n for n in needles if n not in body]
    if missing:
        raise AssertionError(f"missing {missing}")


def _dashboard(page) -> None:
    page.click("text=Sibline Port")
    page.wait_for_selector("text=Progress S-curve", timeout=8000)
    page.wait_for_selector(".chart svg", timeout=8000)
    lines = page.locator(".chart polyline").count()
    if lines < 2:
        raise AssertionError(f"expected planned and earned lines, found {lines}")
    page.wait_for_timeout(400)
    page.screenshot(path=str(SHOTS / "03-dashboard.png"), full_page=True)


def _hover_tooltip(page) -> None:
    page.locator(".chart .hit").nth(20).hover()
    page.wait_for_selector(".chart-tip:not([hidden])", timeout=4000)
    text = page.text_content(".chart-tip")
    if "Planned" not in text:
        raise AssertionError(f"tooltip did not show a series: {text!r}")


def _progress_page(page) -> None:
    page.click("a:has-text('Progress')")
    page.wait_for_selector("text=Progress update", timeout=8000)
    page.wait_for_selector("text=Marine Design")
    page.screenshot(path=str(SHOTS / "04-progress.png"), full_page=True)


def _record_progress(page) -> None:
    """Progress is reported by clicking the row and choosing a workflow step."""
    row = page.locator("tr", has_text="Coastal numerical modelling").first
    before = page.url
    row.locator(".cell-open[data-cell]").first.click()
    page.wait_for_selector("form.cell-form select[name=status_key]", timeout=8000)
    page.select_option("form.cell-form select[name=status_key]", label="Submitted to client — 80%")
    page.wait_for_timeout(1200)

    if page.url != before:
        raise AssertionError("reporting progress should not reload the page")
    text = page.locator("tr", has_text="Coastal numerical modelling").first.text_content()
    if "80%" not in text or "Submitted to client" not in text:
        raise AssertionError(f"status did not persist: {text[:160]}")


def _raise_revision(page) -> None:
    """The client returns comments instead of a Code A."""
    row = page.locator("tr", has_text="Coastal numerical modelling").first
    row.locator("a:has-text('Code B / C')").click()
    page.wait_for_selector("input[name=comments_date]", timeout=8000)
    page.select_option("select[name=code]", "C")
    page.fill("input[name=comments_date]", "05/09/2026")
    page.fill("input[name=note]", "Not approved")
    page.click("button:has-text('Raise revision')")
    page.wait_for_selector("text=Code C", timeout=8000)
    text = page.locator("tr", has_text="Coastal numerical modelling").first.text_content()
    if "Rev 1" not in text:
        raise AssertionError(f"revision not shown: {text[:160]}")
    page.screenshot(path=str(SHOTS / "05-revision.png"), full_page=True)


def _schedule_excel(page) -> None:
    """The dates and durations download, survive an edit in Excel, and import."""
    from openpyxl import load_workbook

    _schedule_page(page, "dates")
    card = _card(page, "Dates and durations")
    with page.expect_download(timeout=10000) as download:
        card.locator("a:has-text('Export to Excel')").click()
    workbook = str(SHOTS / "schedule.xlsx")
    download.value.save_as(workbook)
    if not download.value.suggested_filename.endswith(".xlsx"):
        raise AssertionError("the schedule did not download as a workbook")

    sheet = load_workbook(workbook)["Schedule"]
    wbs = str(sheet.cell(row=2, column=1).value)
    if not wbs or wbs[0].isalpha():
        raise AssertionError(f"the first data row should be a WBS number, got {wbs!r}")

    # Edit it the way anyone would: stretch the first deliverable by a fortnight.
    was = int(sheet.cell(row=2, column=4).value)
    sheet.cell(row=2, column=4).value = was + 14
    sheet.parent.save(workbook)

    card.locator("input[name=workbook]").set_input_files(workbook)
    page.once("dialog", lambda dialog: dialog.accept())
    card.locator("button:has-text('Import')").click()
    page.wait_for_selector("text=rescheduled", timeout=8000)
    if "skipped" in page.text_content("body"):
        raise AssertionError("its own export should import without a complaint")

    row = _card(page, "Dates and durations").locator("tbody tr").first
    if f"{was + 14}d" not in row.inner_text():
        raise AssertionError(f"the table still reads {was}d, not the imported {was + 14}d")
    page.screenshot(path=str(SHOTS / "24-schedule-excel.png"), full_page=True)


def _schedule_reading(page) -> None:
    """The page opens on its charts; a bar says what it is; the diagram marks
    where each path ends and says how many there are."""
    page.goto(f"{BASE}/projects/1/schedule", wait_until="networkidle")

    panels = page.locator("details.panel")
    if panels.count() != 2:
        raise AssertionError(f"expected a panel under each chart, found {panels.count()}")
    if panels.first.get_attribute("open") is not None:
        raise AssertionError("the page should open on the chart, not the table")

    # Hovering a bar — over its WBS label, left of the plotting area — says
    # what the line is about.
    label = page.locator(".chart svg text").filter(has_text="1.1").first
    label.scroll_into_view_if_needed()
    at = label.bounding_box()
    page.mouse.move(at["x"] + at["width"] / 2, at["y"] + at["height"] / 2)
    page.wait_for_selector(".chart-tip:not([hidden])", timeout=4000)
    said = page.locator(".chart-tip:not([hidden])").first.inner_text()
    for expected in ("1.1", "Start", "Duration", "Float"):
        if expected not in said:
            raise AssertionError(f"the bar's tooltip does not say {expected!r}: {said!r}")

    # The toggle opens the table, and the choice is remembered.
    panels.first.locator("summary").click()
    page.wait_for_timeout(400)
    if panels.first.get_attribute("open") is None:
        raise AssertionError("the details toggle did not open the table")
    page.reload(wait_until="networkidle")
    if page.locator("details.panel").first.get_attribute("open") is None:
        raise AssertionError("the browser should remember the panel was left open")

    # Where each path ends, and how many there are.
    page.locator("#network").scroll_into_view_if_needed()
    page.wait_for_timeout(300)
    if page.locator("#network rect[stroke*='series-1']").count() == 0:
        raise AssertionError("a line nothing waits on should be drawn blue")
    body = page.text_content("body")
    for expected in ("unique path", "End of a path", "nothing waits on"):
        if expected not in body:
            raise AssertionError(f"the diagram does not say {expected!r}")
    page.screenshot(path=str(SHOTS / "25-schedule-reading.png"), full_page=True)


def _schedule_simplify(page) -> None:
    """Each kind of link reads differently, and Simplify untangles the picture
    without the page reloading."""
    _schedule_page(page, "links")

    # Wire the diagram up crossed, with one of each kind of link.
    wires = [("11", "1", "FS", "0"), ("10", "2", "SS", "-5"),
             ("9", "3", "FF", "10"), ("8", "4", "SF", "3")]
    for successor, predecessor, kind, lag in wires:
        page.select_option("select[name=successor_id]", value=successor)
        page.select_option("select[name=predecessor_id]", value=predecessor)
        page.select_option("select[name=kind]", kind)
        page.fill("input[name=lag_days]", lag)
        page.click("button:has-text('Link them')")
        page.wait_for_timeout(700)

    # Four colours, four dashes, and a legend naming all four.
    colours = page.eval_on_selector_all(
        "#network path.net-edge", "nodes => nodes.map(n => n.getAttribute('stroke'))")
    if len(set(colours)) < 4:
        raise AssertionError(f"the four kinds should not share a colour: {sorted(set(colours))}")
    dashes = page.eval_on_selector_all(
        "#network path.net-edge",
        "nodes => nodes.map(n => n.getAttribute('stroke-dasharray') || 'solid')")
    if len(set(dashes)) < 4:
        raise AssertionError(f"the four kinds should not share a dash: {sorted(set(dashes))}")
    body = page.text_content("body")
    for named in ("FS · finish → start", "SS · start → start",
                  "FF · finish → finish", "SF · start → finish"):
        if named not in body:
            raise AssertionError(f"the legend does not name {named!r}")

    page.locator("#network").scroll_into_view_if_needed()
    page.wait_for_timeout(300)
    page.screenshot(path=str(SHOTS / "26-links-tangled.png"), full_page=True)

    url = page.url
    page.click("button:has-text('Simplify')")
    page.wait_for_selector(".flash:has-text('Simplif')", timeout=8000)
    said = page.locator(".flash").last.inner_text()
    if page.url != url:
        raise AssertionError("simplifying should not reload the page")
    if "down from" not in said:
        raise AssertionError(f"expected a count of the crossings it removed: {said!r}")

    # And it was written down, so the boxes can be nudged from there.
    page.reload(wait_until="networkidle")
    placed = page.eval_on_selector_all(
        "#network .net-node", "nodes => nodes.map(n => n.dataset.y)")
    if len(set(placed)) < 2:
        raise AssertionError("the new places did not stick")
    page.screenshot(path=str(SHOTS / "27-links-simplified.png"), full_page=True)

    # Tidy up forgets them again.
    page.click("button:has-text('Tidy up')")
    page.wait_for_timeout(1500)
    if page.locator("#network .net-node").count() == 0:
        raise AssertionError("the diagram disappeared")


def _dates_read_dd_mm(page) -> None:
    body = page.text_content("body")
    if "31/08/2026" not in body:
        raise AssertionError("dates should read dd/mm/yyyy")
    if page.locator("input[type=date]").count():
        raise AssertionError("a native date picker would follow the machine's locale")


def _sorting(page) -> None:
    page.click("a:has-text('Progress')")
    page.wait_for_selector("text=Progress update", timeout=8000)
    page.click("th a:has-text('Variance')")
    page.wait_for_selector("text=All deliverables", timeout=8000)
    first = page.locator("tbody tr").first.text_content()
    if "1.6" not in first:
        raise AssertionError(f"worst variance should sort first, got {first[:80]}")
    page.click("th a:has-text('WBS')")
    page.wait_for_selector("text=Sec. 3.1 Marine Design", timeout=8000)
    page.screenshot(path=str(SHOTS / "12-sorted.png"), full_page=True)


def _stepped_planned(page) -> None:
    """A workflow line's planned figure should read one of the step values and
    never something between. Simple lines are pro rata by time, so they are
    allowed any percentage and are excluded here."""
    values = page.eval_on_selector_all(
        "tr[data-tracking='workflow'] td:nth-child(5)",
        "els => els.map(e => e.textContent.trim())",
    )
    if not values:
        raise AssertionError("no workflow rows found to check")
    seen = {v for v in values if v.endswith("%")}
    stray = seen - {"0%", "10%", "40%", "60%", "80%", "100%"}
    if stray:
        raise AssertionError(f"planned showed values between steps: {sorted(stray)}")

    # And the simple lines really are being tracked differently.
    if page.locator("tr[data-tracking='simple']").count() == 0:
        raise AssertionError("expected some lines tracked as a simple percentage")


def _save_all(page) -> None:
    page.fill("input[name='max_revisions']", "8")
    page.click("button:has-text('Save all changes')")
    page.wait_for_selector("text=Saved — project settings", timeout=8000)
    if page.input_value("input[name='max_revisions']") != "8":
        raise AssertionError("the saved value did not come back")


def _print_to_pdf(page) -> None:
    """Each report tab offers a print button and carries a print-only header."""
    for tab, heading in [("Progress", "Progress update"), ("Schedule", "Schedule"),
                         ("Budget", "Budget control"), ("Period", "Period report")]:
        page.click(f"a.tabs >> nth=0" if False else f"nav.tabs a:has-text('{tab}')")
        page.wait_for_selector(f"text={heading}", timeout=8000)
        if page.locator("[data-print]").count() == 0:
            raise AssertionError(f"{tab} has no print button")
        if page.locator(".print-header").count() == 0:
            raise AssertionError(f"{tab} has no print header")

    # Render the page as the printer sees it, which is how a PDF comes out.
    page.emulate_media(media="print")
    page.wait_for_timeout(400)
    hidden = page.evaluate(
        "getComputedStyle(document.querySelector('.topbar')).display === 'none'"
    )
    if not hidden:
        raise AssertionError("the navigation is still on the page when printing")
    page.screenshot(path=str(SHOTS / "13-print.png"), full_page=True)
    page.emulate_media(media="screen")


def _setup_lock(page) -> None:
    page.click("a:has-text('Setup')")
    page.wait_for_selector("text=Project setup", timeout=8000)
    if "Locked" not in page.text_content("body"):
        raise AssertionError("the setup sheet should start locked")

    page.fill("input[name=password]", "2026")
    page.click("button:has-text('Unlock')")
    page.wait_for_selector("text=Setup sheet unlocked", timeout=8000)
    body = page.text_content("body")
    for expected in ("Design workflow", "IDC provided", "Maximum revisions",
                     "Rework days", "Export to Excel", "Import from Excel"):
        if expected not in body:
            raise AssertionError(f"setup is missing {expected!r}")
    page.screenshot(path=str(SHOTS / "10-setup.png"), full_page=True)


def _excel_export(page) -> None:
    with page.expect_download(timeout=10000) as download:
        page.click("a:has-text('Export to Excel')")
    name = download.value.suggested_filename
    if not name.endswith(".xlsx"):
        raise AssertionError(f"unexpected download: {name}")


def _filter_late(page) -> None:
    page.click("a:has-text('Late')")
    page.wait_for_timeout(400)
    if "kick-off" not in page.text_content("body"):
        raise AssertionError("expected the late kick-off milestone")
    page.click("a:has-text('All')")
    page.wait_for_timeout(300)


def _schedule(page) -> None:
    page.click("nav.tabs a:has-text('Schedule')")
    page.wait_for_selector("text=Dates and durations", timeout=8000)
    body = page.text_content("body")
    for expected in ("Programme", "Dependencies", "Duration", "Float"):
        if expected not in body:
            raise AssertionError(f"the schedule is missing {expected!r}")
    for gone in ("Late deliverables", "Behind plan"):
        if gone in body:
            raise AssertionError(f"{gone!r} belongs on the Progress tab now")
    for mark in ("Planned bar", "IDC (workflow only)", "Submission",
                 "Code A due (workflow only)", "Rework after a Code B or C", "Today"):
        if mark not in body:
            raise AssertionError(f"the legend does not say what {mark!r} is")
    if page.locator(".chart svg circle").count() == 0:
        raise AssertionError("the IDC marks are missing from the bars")
    if page.locator(".chart svg polygon").count() == 0:
        raise AssertionError("the submission and Code A stars are missing")
    if "2026" not in page.text_content(".chart"):
        raise AssertionError("the chart should name the months it covers")
    page.screenshot(path=str(SHOTS / "21-schedule.png"), full_page=True)


def _schedule_links(page) -> None:
    """Linking two lines sequences them, and moving one moves the other."""
    _schedule_page(page, "links")
    page.select_option("select[name=successor_id]", index=2)
    page.select_option("select[name=predecessor_id]", index=1)
    page.select_option("select[name=kind]", "FS")
    page.click("button:has-text('Link them')")
    page.wait_for_selector("text=Dependency added", timeout=8000)
    if page.locator("text=On the critical path").count() == 0:
        raise AssertionError("the network should name the critical path")
    if page.locator("svg path[marker-end]").count() == 0:
        raise AssertionError("the dependency arrows are missing")
    page.screenshot(path=str(SHOTS / "22-network.png"), full_page=True)


def _schedule_deps(page) -> None:
    """A lag and a link type change in the row; a box can be dragged; a
    dependency is removed without the page reloading."""
    _schedule_page(page, "links")
    # A second link, this one start-to-start.
    page.select_option("select[name=successor_id]", index=3)
    page.select_option("select[name=predecessor_id]", index=1)
    page.select_option("select[name=kind]", "SS")
    page.fill("input[name=lag_days]", "-5")          # work that overlaps
    page.click("button:has-text('Link them')")
    page.wait_for_selector("text=Dependency added", timeout=8000)
    if page.locator("svg path[stroke-dasharray]").count() == 0:
        raise AssertionError("a start-to-start link should draw differently")
    if "-5d" not in page.text_content("body"):
        raise AssertionError("a negative lag should be kept")

    url = page.url
    page.locator("[data-cell='link-lag']").first.click()
    page.wait_for_selector("form.cell-form input[name=lag_days]", timeout=6000)
    page.fill("form.cell-form input[name=lag_days]", "12")
    page.locator("form.cell-form input[name=lag_days]").blur()
    page.wait_for_timeout(1400)
    if page.url != url:
        raise AssertionError("changing a lag should not reload the page")
    if "12d" not in page.locator("[data-cell='link-lag']").first.inner_text():
        raise AssertionError("the lag did not stick")

    page.locator("[data-cell='link-kind']").first.click()
    page.wait_for_selector("form.cell-form select[name=kind]", timeout=6000)
    page.select_option("form.cell-form select[name=kind]", "SS")
    page.wait_for_timeout(1400)
    if "start → start" not in page.locator("[data-cell='link-kind']").first.inner_text():
        raise AssertionError("the link type did not stick")

    # Either end of the link moves in the row too, and the row redraws with the
    # deliverable it now points at.
    link = page.locator("tr[id^='link-']").first.get_attribute("id")
    end = page.locator(f"#{link} [data-cell='link-successor']")
    waits_on = page.locator(f"#{link} [data-cell='link-predecessor']").get_attribute("data-value")
    before = end.inner_text().strip()
    end.click()
    page.wait_for_selector("form.cell-form select[name=successor_id]", timeout=6000)
    choices = page.locator("form.cell-form select[name=successor_id] option")
    moved_to = next(                       # a late line: nothing yet waits on it
        choices.nth(i).get_attribute("value")
        for i in reversed(range(choices.count()))
        if choices.nth(i).get_attribute("value") != waits_on
    )
    page.select_option("form.cell-form select[name=successor_id]", moved_to)
    page.wait_for_timeout(1400)
    if page.url != url:
        raise AssertionError("moving an end of a link should not reload the page")
    end = page.locator(f"#{link} [data-cell='link-successor']")
    if end.inner_text().strip() == before:
        raise AssertionError(f"the link still waits on {before!r}")
    if end.get_attribute("data-value") != moved_to:
        raise AssertionError("the cell did not take the deliverable it now points at")

    # A link onto itself is refused, and says so without losing the page.
    end.click()
    page.wait_for_selector("form.cell-form select[name=successor_id]", timeout=6000)
    page.select_option("form.cell-form select[name=successor_id]", waits_on)
    page.wait_for_selector(".flash.error", timeout=6000)
    if page.url != url:
        raise AssertionError("a refused change should not throw the page away")
    if page.locator(f"#{link} [data-cell='link-successor']").get_attribute("data-value") != moved_to:
        raise AssertionError("a refused change should leave the cell as it was")

    page.reload(wait_until="networkidle")
    if page.locator(f"#{link} [data-cell='link-successor']").inner_text().strip() == before:
        raise AssertionError("the moved end did not survive a reload")

    # A box can be dragged out of the way, and stays there.
    box = page.locator(".net-node.movable").first
    box.scroll_into_view_if_needed()
    page.wait_for_timeout(300)
    before = (box.get_attribute("data-x"), box.get_attribute("data-y"))
    place = box.bounding_box()
    page.mouse.move(place["x"] + place["width"] / 2, place["y"] + place["height"] / 2)
    page.mouse.down()
    page.mouse.move(place["x"] + place["width"] / 2 + 160,
                    place["y"] + place["height"] / 2 + 90, steps=10)
    page.mouse.up()
    page.wait_for_timeout(1200)

    moved = page.locator(".net-node.movable").first
    if (moved.get_attribute("data-x"), moved.get_attribute("data-y")) == before:
        raise AssertionError("the box did not move")
    page.reload(wait_until="networkidle")
    kept = page.locator(".net-node.movable").first
    if (kept.get_attribute("data-x"), kept.get_attribute("data-y")) == before:
        raise AssertionError("the box did not stay where it was put")
    page.screenshot(path=str(SHOTS / "23-dependencies.png"), full_page=True)

    # Removing one takes its row with it, and nothing reloads.
    rows = page.locator("tr[id^='link-']").count()
    page.once("dialog", lambda dialog: dialog.accept())
    url = page.url
    page.locator("form[data-live-remove] button").first.click()
    page.wait_for_timeout(1400)
    if page.url != url:
        raise AssertionError("removing a dependency should not reload the page")
    if page.locator("tr[id^='link-']").count() != rows - 1:
        raise AssertionError("the row was not removed")

    # The dependencies go out to Excel and come back.
    card = _card(page, "Dependencies")
    with page.expect_download(timeout=10000) as download:
        card.locator("a:has-text('Export to Excel')").click()
    workbook = str(SHOTS / "dependencies.xlsx")
    download.value.save_as(workbook)
    if not download.value.suggested_filename.endswith(".xlsx"):
        raise AssertionError("the dependencies did not download as a workbook")

    card.locator("input[name=workbook]").set_input_files(workbook)
    page.once("dialog", lambda dialog: dialog.accept())
    card.locator("button:has-text('Import')").click()
    page.wait_for_selector("text=imported", timeout=8000)
    if "skipped" in page.text_content("body"):
        raise AssertionError("its own export should import without a complaint")

    # And Tidy up puts the boxes back under the automatic layout.
    page.click("button:has-text('Tidy up')")
    page.wait_for_timeout(1400)
    tidied = page.locator(".net-node.movable").first
    if (tidied.get_attribute("data-x"), tidied.get_attribute("data-y")) != ("16", "16"):
        raise AssertionError("tidy up should return the boxes to the layout")


def _schedule_amend(page) -> None:
    _schedule_page(page, "dates")
    before = page.locator("#submission-1").inner_text().strip()
    url = page.url
    page.locator("#duration-1 .cell-open").click()
    page.wait_for_selector("form.cell-form input[name=duration_days]", timeout=8000)
    page.fill("form.cell-form input[name=duration_days]", "45")
    page.locator("form.cell-form input[name=duration_days]").blur()
    page.wait_for_timeout(1600)

    if page.url != url:
        raise AssertionError("amending the plan should not reload the page")
    after = page.locator("#submission-1").inner_text().strip()
    if after == before:
        raise AssertionError(f"the finish did not follow the duration: {before} -> {after}")
    if page.locator("#duration-1 .cell-open").count() != 1:
        raise AssertionError("the cell must stay clickable after a change")
    if "45d" not in page.locator("#duration-1").inner_text():
        raise AssertionError("the duration did not stick")


def _budget(page) -> None:
    page.click("a:has-text('Budget')")
    page.wait_for_selector("text=Budget control", timeout=8000)
    page.wait_for_selector(".chart svg", timeout=8000)
    page.screenshot(path=str(SHOTS / "06-budget.png"), full_page=True)


def _book_hours(page) -> None:
    page.click("a:has-text('Timesheet')")
    page.wait_for_selector("text=Book hours", timeout=8000)
    page.select_option("select[name=trade_id]", label="Geotechnical")
    page.fill("input[name=hours]", "36")
    page.fill("input[name=description]", "Borehole data review")
    page.click("button:has-text('Book hours')")
    page.wait_for_selector("text=Booked 36 hours", timeout=8000)
    if "Borehole data review" not in page.text_content("body"):
        raise AssertionError("entry not listed")
    page.screenshot(path=str(SHOTS / "07-timesheet.png"), full_page=True)

    page.click("a:has-text('Budget')")
    page.wait_for_selector("text=Budget control", timeout=8000)
    if "36 h" not in page.text_content("body"):
        raise AssertionError("booked hours did not reach budget control")


def _period(page) -> None:
    page.click("a:has-text('Period')")
    page.wait_for_selector("text=Period report", timeout=8000)
    page.wait_for_selector("text=Earned in period by trade")
    page.screenshot(path=str(SHOTS / "08-period.png"), full_page=True)


def _minutes_capture(page) -> None:
    """The roster is typed once, then ticked; items carry an owner and an impact."""
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("text=Attendance list", timeout=8000)

    for name, org, role in [("Ahmed Mitwally", "Dar", "Project manager"),
                            ("Client Rep", "Sibline Port Authority", "Design manager")]:
        roster = page.locator("form", has=page.locator("button:has-text('Add attendee')"))
        roster.locator("input[name=name]").fill(name)
        roster.locator("input[name=organisation]").fill(org)
        roster.locator("input[name=job_title]").fill(role)
        roster.locator("button:has-text('Add attendee')").click()
        page.wait_for_selector(f"text={name} added to the attendance list", timeout=8000)

    meeting_form = page.locator("form", has=page.locator("button:has-text('Add meeting')"))
    meeting_form.locator("input[name=ref]").fill("MOM-01")
    meeting_form.locator("input[name=title]").fill("Weekly design coordination")
    meeting_form.locator("input[name=meeting_date]").fill("03/09/2026")
    meeting_form.locator("input[name=location]").fill("Site office")
    meeting_form.locator("button:has-text('Add meeting')").click()
    page.wait_for_selector("text=tick who attended", timeout=8000)

    # Everyone is invited by default; untick one to show apologies.
    boxes = page.locator("input[name=present]")
    if boxes.count() != 2:
        raise AssertionError(f"expected a tick box per attendee, found {boxes.count()}")
    boxes.nth(1).uncheck()
    page.fill("input[name=chaired_by]", "Ahmed Mitwally")
    page.fill("input[name=next_date]", "10/09/2026")
    page.click("button:has-text('Save meeting')")
    page.wait_for_selector("text=Meeting saved", timeout=8000)
    if "Apologies" not in page.text_content("body"):
        raise AssertionError("the attendee who was unticked should show apologies")

    for subject, agreement, impact, due, owner in [
        ("Quay wall levels", "Marine to reissue the layout", "Time", "10/09/2026", "MR"),
        ("Additional bathymetric survey", "Client to confirm the budget", "Cost", "01/08/2026", "Client"),
    ]:
        form = page.locator("form", has=page.locator("button:has-text('Add item')"))
        form.locator("input[name=subject]").fill(subject)
        form.locator("textarea[name=agreement]").fill(agreement)
        form.locator("select[name=impact]").select_option(label=impact)
        form.locator("input[name=due_date]").fill(due)
        form.locator("select[name=owner_code]").select_option(owner)
        # An item can bear on more than one trade at a time.
        form.locator("input[name=trade_ids]").first.check()
        form.locator("input[name=trade_ids]").nth(1).check()
        form.locator("button:has-text('Add item')").click()
        page.wait_for_selector("text=Item added", timeout=8000)

    body = page.text_content("body")
    for expected in ("Quay wall levels", "Marine to reissue the layout", "SIBLINE-PORT",
                     "MR", "Marine", "Geotechnical"):
        if expected not in body:
            raise AssertionError(f"the minutes are missing {expected!r}")
    page.screenshot(path=str(SHOTS / "14-meeting.png"), full_page=True)


def _minutes_filters(page) -> None:
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("h1:has-text('Minutes of meeting')", timeout=8000)
    body = page.text_content("body")
    if "Quay wall levels" not in body:
        raise AssertionError("open items should be shown by default")
    if "MOM-01" not in body:
        raise AssertionError("the meeting reference did not save")

    page.click(".chips a:has-text('Overdue')")
    page.wait_for_timeout(400)
    body = page.text_content("body")
    if "Additional bathymetric survey" not in body or "Quay wall levels" in body:
        raise AssertionError("the overdue filter did not narrow the register")

    page.click(".chips a:has-text('All items')")
    page.wait_for_timeout(300)
    page.fill("input[name=q]", "quay")
    page.click("button:has-text('Apply')")
    page.wait_for_timeout(400)
    body = page.text_content("body")
    if "Quay wall levels" not in body or "bathymetric" in body:
        raise AssertionError("the keyword search did not narrow the register")

    page.click("a:has-text('Clear')")
    page.wait_for_timeout(300)
    page.click("th a:has-text('Due')")
    page.wait_for_timeout(400)
    page.screenshot(path=str(SHOTS / "15-minutes.png"), full_page=True)


def _minutes_word(page) -> None:
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("h1:has-text('Minutes of meeting')", timeout=8000)
    with page.expect_download(timeout=10000) as download:
        page.click("a:has-text('Export Word')")
    name = download.value.suggested_filename
    if not name.endswith(".docx"):
        raise AssertionError(f"unexpected download: {name}")

    page.click("table a:has-text('Weekly design coordination')")
    page.wait_for_selector("text=Items and agreements", timeout=8000)
    with page.expect_download(timeout=10000) as download:
        page.click("a:has-text('Export Word')")
    if not download.value.suggested_filename.endswith(".docx"):
        raise AssertionError("the minutes did not download as a Word document")


def _minutes_reorder(page) -> None:
    """An item's number is its position, so moving it renumbers both rows."""
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("h1:has-text('Minutes of meeting')", timeout=8000)
    page.click("table a:has-text('Weekly design coordination')")
    page.wait_for_selector("text=Items and agreements", timeout=8000)

    rows = page.locator("tbody tr:has(button[title='Move down'])")
    before = [rows.nth(i).text_content() for i in range(rows.count())]
    if len(before) != 2 or "Quay wall levels" not in before[0]:
        raise AssertionError(f"unexpected starting order: {before}")
    if "1.1" not in before[0] or "1.2" not in before[1]:
        raise AssertionError("items should start numbered 1.1 and 1.2")

    # The first item cannot go up and the last cannot go down.
    if not page.locator("button[title='Move up']").first.is_disabled():
        raise AssertionError("the first item should not be movable up")
    if not page.locator("button[title='Move down']").last.is_disabled():
        raise AssertionError("the last item should not be movable down")

    page.locator("button[title='Move down']").first.click()
    page.wait_for_selector("text=Items and agreements", timeout=8000)
    rows = page.locator("tbody tr:has(button[title='Move down'])")
    after = [rows.nth(i).text_content() for i in range(rows.count())]
    if "Additional bathymetric survey" not in after[0] or "1.1" not in after[0]:
        raise AssertionError(f"the moved item did not take number 1.1: {after}")
    if "Quay wall levels" not in after[1] or "1.2" not in after[1]:
        raise AssertionError(f"the swapped item did not take number 1.2: {after}")
    page.screenshot(path=str(SHOTS / "17-reordered.png"), full_page=True)

    page.locator("button[title='Move up']").last.click()      # put it back
    page.wait_for_selector("text=Items and agreements", timeout=8000)


def _minutes_edit_in_place(page) -> None:
    """Edit opens the form where it stands — no round trip, no jump."""
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("h1:has-text('Minutes of meeting')", timeout=8000)
    page.click("table a:has-text('Weekly design coordination')")
    page.wait_for_selector("text=Items and agreements", timeout=8000)

    row = page.locator("tr", has_text="Quay wall levels").first
    item_id = row.get_attribute("id").split("-")[1]
    editor = page.locator(f"#edit-{item_id}")
    if editor.is_visible():
        raise AssertionError("the edit form should start closed")

    url_before = page.url
    row.locator("a:has-text('Edit')").click()
    editor.wait_for(state="visible", timeout=4000)
    if page.url != url_before:
        raise AssertionError("editing should not navigate away from the page")

    box = editor.locator("textarea[name=agreement]")
    if not box.is_visible():
        raise AssertionError("the agreement should be a text area with room to write")
    if box.bounding_box()["height"] < 60:
        raise AssertionError("the writing box is too small to read back before saving")

    box.fill("Marine to reissue the layout with the revised levels agreed today")
    editor.locator("button:has-text('Save item')").click()
    page.wait_for_selector("text=Item saved", timeout=8000)
    if "revised levels agreed today" not in page.text_content("body"):
        raise AssertionError("the edited agreement did not save")
    page.screenshot(path=str(SHOTS / "18-edit-in-place.png"), full_page=True)


def _minutes_cells(page) -> None:
    """Owner, trades, affects and the date are changed in the row itself."""
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("h1:has-text('Minutes of meeting')", timeout=8000)
    row = page.locator("tr", has_text="Quay wall levels").first
    item_id = row.get_attribute("id").split("-")[1]
    url_before = page.url

    # Owner: click the cell, pick another party.
    row.locator("[data-cell='owner']").click()
    page.wait_for_selector(f"#owner-{item_id} select", timeout=4000)
    page.select_option(f"#owner-{item_id} select", "PMC")
    page.wait_for_selector(f"#owner-{item_id} .cell-open:has-text('PMC')", timeout=6000)
    if page.url != url_before:
        raise AssertionError("changing a field should not reload the page")

    # Affects: the same, and the row's wording follows.
    row.locator("[data-cell='impact']").click()
    page.wait_for_selector(f"#impact-{item_id} select", timeout=4000)
    page.select_option(f"#impact-{item_id} select", "both")
    page.wait_for_selector(f"#impact-{item_id} .cell-open:has-text('Time & cost')", timeout=6000)

    # Trades: tick another one; the cell lists them all.
    row.locator("[data-cell='trades']").click()
    page.wait_for_selector(f"#trades-{item_id} input[type=checkbox]", timeout=4000)
    page.locator(f"#trades-{item_id} input[type=checkbox]").nth(2).check()
    page.wait_for_timeout(900)
    listed = page.text_content(f"#trades-{item_id}")
    if "Marine Structures" not in listed:
        raise AssertionError(f"the trade cell did not take the new trade: {listed!r}")

    # The date, and the status badge that reads on it.
    row.locator("[data-cell='due']").click()
    page.wait_for_selector(f"#due-{item_id} input", timeout=4000)
    page.fill(f"#due-{item_id} input", "01/01/2020")
    page.locator(f"#due-{item_id} input").blur()
    page.wait_for_selector(f"#status-{item_id}:has-text('overdue')", timeout=6000)

    page.screenshot(path=str(SHOTS / "20-cells.png"), full_page=True)

    # And it really was written, not just drawn.
    page.reload(wait_until="networkidle")
    body = page.text_content("body")
    for expected in ("PMC", "Time & cost", "Marine Structures", "01/01/2020"):
        if expected not in body:
            raise AssertionError(f"{expected!r} did not survive a reload")


def _minutes_calendar(page) -> None:
    """Clicking a date field opens a calendar that writes back dd/mm/yyyy."""
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("h1:has-text('Minutes of meeting')", timeout=8000)

    field = page.locator("form input[name=due_date]").first
    field.click()
    page.wait_for_selector(".calendar", state="visible", timeout=4000)

    month = page.text_content(".cal-month")
    page.click(".cal-nav[data-step='1']")
    if page.text_content(".cal-month") == month:
        raise AssertionError("the next-month arrow did not move the calendar")
    page.click(".cal-nav[data-step='-1']")

    page.locator(".cal-day:not(.outside)").nth(14).click()
    page.wait_for_selector(".calendar", state="hidden", timeout=4000)
    value = field.input_value()
    if not re.fullmatch(r"\d{2}/\d{2}/\d{4}", value):
        raise AssertionError(f"the calendar wrote {value!r}, not dd/mm/yyyy")
    if not value.startswith("15/"):
        raise AssertionError(f"the fifteenth of the month should give 15/..., got {value}")

    # Typing still works, and the field is not a native picker.
    field.fill("07/10/2026")
    if page.locator("input[type=date]").count():
        raise AssertionError("a native date picker would follow the machine's locale")
    page.screenshot(path=str(SHOTS / "19-calendar.png"), full_page=True)


def _minutes_agenda(page) -> None:
    page.click("nav.tabs a:has-text('Minutes')")
    page.wait_for_selector("h1:has-text('Minutes of meeting')", timeout=8000)
    page.click("a:has-text('Next-meeting agenda')")
    page.wait_for_selector("text=Agenda", timeout=8000)
    body = page.text_content("body")
    for expected in ("MR", "Client", "Quay wall levels", "Additional bathymetric survey"):
        if expected not in body:
            raise AssertionError(f"the agenda is missing {expected!r}")
    page.screenshot(path=str(SHOTS / "16-agenda.png"), full_page=True)


def _dark(page) -> None:
    page.click("a:has-text('Dashboard')")
    page.wait_for_selector("text=Progress S-curve", timeout=8000)
    page.evaluate("localStorage.setItem('pm-theme','dark');document.documentElement.setAttribute('data-theme','dark')")
    page.wait_for_timeout(500)
    page.screenshot(path=str(SHOTS / "10-dashboard-dark.png"), full_page=True)


def _live(page) -> None:
    """A second window picks up a change on its own, without a refresh."""
    other = page.context.browser.new_context()
    watcher = other.new_page()
    watcher.goto(f"{BASE}/login", wait_until="networkidle")
    watcher.fill("input[name=email]", EMAIL)
    watcher.fill("input[name=password]", PASSWORD)
    watcher.click("button[type=submit]")
    watcher.wait_for_selector("text=Portfolio", timeout=8000)
    watcher.goto(f"{BASE}/projects/1/schedule?panel=dates", wait_until="networkidle")
    before = watcher.locator("#duration-1").inner_text().strip()

    _schedule_page(page, "dates")
    page.locator("#duration-1 .cell-open").click()
    page.wait_for_selector("form.cell-form input[name=duration_days]", timeout=8000)
    page.fill("form.cell-form input[name=duration_days]", "28")
    page.locator("form.cell-form input[name=duration_days]").blur()
    page.wait_for_timeout(1500)

    watcher.wait_for_timeout(20000)          # the live check runs every 15 seconds
    after = watcher.locator("#duration-1").inner_text().strip()
    other.close()
    if after == before:
        raise AssertionError(f"the other window did not pick the change up: {before} -> {after}")


def _mobile(page) -> None:
    page.set_viewport_size({"width": 390, "height": 844})
    page.wait_for_timeout(600)
    overflow = page.evaluate(
        "document.documentElement.scrollWidth - document.documentElement.clientWidth"
    )
    if overflow > 2:
        raise AssertionError(f"page scrolls horizontally by {overflow}px")
    page.screenshot(path=str(SHOTS / "11-mobile.png"), full_page=True)


if __name__ == "__main__":
    raise SystemExit(main())
